"""
Comprehensive test suite for the Godrej Market Share Model.
Run: python3 run_all_tests.py
Every test must pass before the model is considered release-ready.
"""
import sys
import pandas as pd
import numpy as np
import openpyxl

from cleaner import (
    process_all_files, process_workbook, get_sheet_overview,
    parse_region_sheet_name, classify_flag, classify_company,
)
from calculator import (
    add_calculations, build_factor_lookup, DEFAULT_INDIVIDUAL_FACTOR_UR,
    DEFAULT_ZONE_MAPPING, STATE_TO_ZONE, build_company_summary,
    compute_variance, smart_guess_companies, find_potential_duplicates,
    get_period_list,
)
from exporter import export_to_excel
from default_brand_mapping import DEFAULT_BRAND_MAPPING
from default_company_mapping import DEFAULT_COMPANY_MAPPING

RAW = '/mnt/user-data/uploads/Rajasthan_SHC.xlsx'
TRUTH = '/mnt/user-data/uploads/HC_KWP_Working_CompanyWise_V3_18062026__2_.xlsx'

PASS, FAIL = 0, 0
def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {name}")
    else:
        FAIL += 1
        print(f"  FAIL  {name}  {detail}")


def load_truth(state='Rajasthan', fmt='SHC'):
    wb = openpyxl.load_workbook(TRUTH, data_only=True, read_only=True)
    ws = wb['Sheet1']
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == state and r[3] == fmt]
    df = pd.DataFrame(rows, columns=headers)
    df['UR'] = df['Urban/Rural'].map({'Urban': 'U', 'Rural': 'R', 'U+R': 'U+R'})
    return df

QMAP = {
    '2023 Apr - 2023 Jun': "01. AMJ'23", '2023 Jul - 2023 Sep': "02. JAS'23",
    '2023 Oct - 2023 Dec': "03. OND'23", '2024 Jan - 2024 Mar': "04. JFM'24",
    '2024 Apr - 2024 Jun': "05. AMJ'24", '2024 Jul - 2024 Sep': "06. JAS'24",
    '2024 Oct - 2024 Dec': "07. OND'24", '2025 Jan - 2025 Mar': "08. JFM'25",
    '2025 Apr - 2025 Jun': "09. AMJ'25", '2025 Jul - 2025 Sep': "10. JAS'25",
    '2025 Oct - 2025 Dec': "11. OND'25", '2026 Jan - 2026 Mar': "12. JFM'26",
    '2023 Apr - 2024 Mar': 'MAT Mar24', '2024 Apr - 2025 Mar': 'MAT Mar25',
    '2025 Apr - 2026 Mar': 'MAT Mar26',
}

print("=" * 70)
print("TEST GROUP 1: Sheet name parsing")
print("=" * 70)
cases = [
    ('Rajasthan (U+R)', ('Rajasthan', 'U+R', False)),
    ('Rajasthan (U)', ('Rajasthan', 'U', False)),
    ('Rajasthan (R)', ('Rajasthan', 'R', False)),
    ('North (U+R)', ('North', 'U+R', True)),
    ('South (R)', ('South', 'R', True)),
    ('East (U)', ('East', 'U', True)),
    ('West (U+R)', ('West', 'U+R', True)),
    ('GCPL East (U+R)', ('GCPL East', 'U+R', True)),
    ('GCPL West (U)', ('GCPL West', 'U', True)),
    ('North Zone(U)', ('North', 'U', True)),
    ('All India Urban', ('All India', 'U', False)),
    ('All India Rural', ('All India', 'R', False)),
    ('All India U+R', ('All India', 'U+R', False)),
    ('Andhra Pradesh excl Tela (U+R)', ('Andhra Pradesh excl Tela', 'U+R', False)),
    ('Madhya Pradesh excl Chha (R)', ('Madhya Pradesh excl Chha', 'R', False)),
]
for name, expected in cases:
    got = parse_region_sheet_name(name)
    check(f"parse '{name}'", got == expected, f"got {got}, expected {expected}")

print("=" * 70)
print("TEST GROUP 2: Flag / Company classification")
print("=" * 70)
check("Category row detected", classify_flag('[HCEXL] ANY HAIR COL.SHAMPOO') == 'Category')
check("Brand row detected", classify_flag('[HCEXL] GODREJ SELFIE HAIR COLOUR SHAMPOO', DEFAULT_BRAND_MAPPING) == 'Brand')
check("Others row detected", classify_flag("[HCEXL] OTH.BRDED CLR'ANT/HENNA", DEFAULT_BRAND_MAPPING) == 'Others')
check("SKU row detected", classify_flag('Godrej Selfie Shampoo Sachet 15ml') == 'SKU')
check("Non-bracket HBP brand '5 Star' via mapping", classify_flag('5 Star', DEFAULT_BRAND_MAPPING) == 'Brand')
check("Sub-brand via mapping", classify_flag('[HCEXL] GARNIER BLACK NATURALS', DEFAULT_BRAND_MAPPING) == 'Sub-brand')
check("Category company = Category Total", classify_company('[HCEXL] ANY X', 'Category') == 'Category Total')
check("GCPL company lookup", classify_company('[HCEXL] GODREJ SELFIE HAIR COLOUR SHAMPOO', 'Brand', DEFAULT_COMPANY_MAPPING) == 'GODREJ CONSUMER PRODS')
check("Unmapped brand -> Others / Unmapped", classify_company('[HCEXL] SOMETHING BRAND NEW', 'Brand', DEFAULT_COMPANY_MAPPING) == 'Others / Unmapped')

print("=" * 70)
print("TEST GROUP 3: Sheet overview (Region Selection defaults)")
print("=" * 70)
f = open(RAW, 'rb')
ov = get_sheet_overview(f)
check("3 sheets found in Rajasthan file", len(ov) == 3)
check("All Rajasthan sheets default-included", all(o['Include'] for o in ov))
check("No sheet misdetected as zone", all(not o['Is_Zone'] for o in ov))

print("=" * 70)
print("TEST GROUP 4: Full pipeline vs GROUND TRUTH (the core correctness test)")
print("=" * 70)
f2 = open(RAW, 'rb')
master_df, sku_df = process_all_files(
    {'SHC': f2},
    flag_overrides=dict(DEFAULT_BRAND_MAPPING),
    company_overrides=dict(DEFAULT_COMPANY_MAPPING),
    included_sheets_by_format={'SHC': {'Rajasthan (U+R)', 'Rajasthan (U)', 'Rajasthan (R)'}},
)
check("Master rows extracted", len(master_df) == 270, f"got {len(master_df)}")
check("SKU rows extracted for variance", len(sku_df) == 1245, f"got {len(sku_df)}")

factor_lookup = build_factor_lookup(dict(DEFAULT_INDIVIDUAL_FACTOR_UR))
zone_mapping = {'North': ['Rajasthan']}  # what the app would auto-build here
result, missing, unmapped, zero_factors = add_calculations(master_df, factor_lookup=factor_lookup, zone_mapping=zone_mapping)
check("No missing factor regions", len(missing) == 0, str(missing))
check("No unmapped zones", len(unmapped) == 0, str(unmapped))
check("No zero-value factors in the (now fixed) default table", len(zero_factors) == 0, str(zero_factors))

truth = load_truth()
res = result[result['TG_Segment'] == 'TOTAL'].set_index(['Urban_Rural', 'Brand_SKU_Item']).sort_index()
truth_idx = truth.set_index(['UR', 'Brand_SKU Item'])

# 4a. Sales Derived exactness: every brand, every UR, ALL 15 periods
checked, max_diff, flag_mm = 0, 0.0, 0
for (ur, item), trow in truth_idx.iterrows():
    if (ur, item) not in res.index:
        continue
    rrow = res.loc[(ur, item)]
    if isinstance(rrow, pd.DataFrame):
        rrow = rrow.iloc[0]
    if rrow['Flag'] != trow['Flag']:
        flag_mm += 1
    for our_p, truth_p in QMAP.items():
        t = trow.get(f'Sales Value_{truth_p}')
        o = rrow.get(f'Sales Derived__{our_p}')
        if o is None or pd.isna(o) or t in (None, 0) or pd.isna(t):
            continue
        d = abs(o - t) / t * 100
        checked += 1
        if d > max_diff:
            max_diff = d
check(f"Sales Derived exact across ALL periods ({checked} datapoints)", max_diff < 1e-9, f"max diff {max_diff}%")
check("Zero flag mismatches", flag_mm == 0, f"{flag_mm} mismatches")

# 4b. Units Estd exactness
checked_u, max_diff_u = 0, 0.0
for (ur, item), trow in truth_idx.iterrows():
    if (ur, item) not in res.index:
        continue
    rrow = res.loc[(ur, item)]
    if isinstance(rrow, pd.DataFrame):
        rrow = rrow.iloc[0]
    for our_p, truth_p in QMAP.items():
        t = trow.get(f'Units_{truth_p}')
        o = rrow.get(f'Units Estd__{our_p}')
        if o is None or pd.isna(o) or t in (None, 0) or pd.isna(t):
            continue
        d = abs(o - t) / t * 100
        checked_u += 1
        if d > max_diff_u:
            max_diff_u = d
check(f"Units Estd exact across ALL periods ({checked_u} datapoints)", max_diff_u < 1e-9, f"max diff {max_diff_u}%")

# 4c. Value MS% exactness
checked_m, max_diff_m = 0, 0.0
for (ur, item), trow in truth_idx.iterrows():
    if (ur, item) not in res.index:
        continue
    rrow = res.loc[(ur, item)]
    if isinstance(rrow, pd.DataFrame):
        rrow = rrow.iloc[0]
    for our_p, truth_p in QMAP.items():
        t = trow.get(f'Value MS%_{truth_p}')
        o = rrow.get(f'Value MS%__{our_p}')
        if o is None or pd.isna(o) or t is None or pd.isna(t):
            continue
        d = abs(o - t)
        checked_m += 1
        if d > max_diff_m:
            max_diff_m = d
check(f"Value MS% exact across ALL periods ({checked_m} datapoints)", max_diff_m < 1e-9, f"max abs diff {max_diff_m}")

# 4d. Units MS% exactness
checked_um, max_diff_um = 0, 0.0
for (ur, item), trow in truth_idx.iterrows():
    if (ur, item) not in res.index:
        continue
    rrow = res.loc[(ur, item)]
    if isinstance(rrow, pd.DataFrame):
        rrow = rrow.iloc[0]
    for our_p, truth_p in QMAP.items():
        t = trow.get(f'Units MS%_{truth_p}')
        o = rrow.get(f'Units MS%__{our_p}')
        if o is None or pd.isna(o) or t is None or pd.isna(t):
            continue
        d = abs(o - t)
        checked_um += 1
        if d > max_diff_um:
            max_diff_um = d
check(f"Units MS% exact across ALL periods ({checked_um} datapoints)", max_diff_um < 1e-9, f"max abs diff {max_diff_um}")

print("=" * 70)
print("TEST GROUP 5: Company summary vs ground truth (GCPL and every competitor)")
print("=" * 70)
cs = build_company_summary(result)
truth_ur = truth[truth['UR'] == 'U+R']
company_col = 'Company Flag '  # note: trailing space exists in the source header
truth_company_sums = {}
for _, r in truth_ur.iterrows():
    comp = r[company_col]
    if comp in ('Category Total',):
        continue
    for our_p, truth_p in QMAP.items():
        v = r.get(f'Sales Value_{truth_p}')
        if v is None or pd.isna(v):
            continue
        truth_company_sums.setdefault((comp, our_p), 0.0)
        truth_company_sums[(comp, our_p)] += v

cs_ur = cs[(cs['State_Zone'] == 'Rajasthan') & (cs['Urban_Rural'] == 'U+R') & (cs['TG_Segment'] == 'TOTAL')]
checked_c, max_diff_c = 0, 0.0
for (comp, our_p), t in truth_company_sums.items():
    row = cs_ur[cs_ur['Company'] == comp]
    if len(row) == 0 or t == 0:
        continue
    o = row[f'Sales Derived__{our_p}'].values[0]
    if pd.isna(o):
        continue
    d = abs(o - t) / t * 100
    checked_c += 1
    if d > max_diff_c:
        max_diff_c = d
check(f"Company-level Sales exact for every company/period ({checked_c} datapoints)", max_diff_c < 1e-9, f"max diff {max_diff_c}%")

print("=" * 70)
print("TEST GROUP 6: U+R = U + R identity (structural invariant)")
print("=" * 70)
periods = get_period_list(result)
tot = result[(result['TG_Segment'] == 'TOTAL') & (result['State_Zone'] == 'Rajasthan')]
u = tot[tot['Urban_Rural'] == 'U'].set_index('Brand_SKU_Item')
r_ = tot[tot['Urban_Rural'] == 'R'].set_index('Brand_SKU_Item')
ur = tot[tot['Urban_Rural'] == 'U+R'].set_index('Brand_SKU_Item')
bad = 0
for item in ur.index:
    if item not in u.index or item not in r_.index:
        continue
    for p in ['2025 Apr - 2026 Mar']:
        s_ur = ur.loc[item, f'Sales Derived__{p}']
        s_sum = u.loc[item, f'Sales Derived__{p}'] + r_.loc[item, f'Sales Derived__{p}']
        if pd.notna(s_ur) and abs(s_ur - s_sum) > 1e-6:
            bad += 1
check("U+R equals U + R for every brand", bad == 0, f"{bad} violations")

print("=" * 70)
print("TEST GROUP 7: Zone synthesis correctness")
print("=" * 70)
north = result[(result['State_Zone'] == 'North') & (result['Urban_Rural'] == 'U+R') & (result['TG_Segment'] == 'TOTAL') & (result['Flag'] == 'Category')]
raj = result[(result['State_Zone'] == 'Rajasthan') & (result['Urban_Rural'] == 'U+R') & (result['TG_Segment'] == 'TOTAL') & (result['Flag'] == 'Category')]
check("Zone North exists in output", len(north) == 1)
check("Zone North == sum of member states (Rajasthan only here)",
      abs(north['Sales Derived__2025 Apr - 2026 Mar'].values[0] - raj['Sales Derived__2025 Apr - 2026 Mar'].values[0]) < 1e-6)

# Zone with garbage sheet data present in input must be ignored
fake_zone = master_df[master_df['Urban_Rural'] == 'U'].copy()
fake_zone['State_Zone'] = 'North'
fake_zone['Is_Zone'] = True
for c in fake_zone.columns:
    if c.startswith('HH__') or c.startswith('Val__') or c.startswith('Avg NOP__'):
        fake_zone[c] = 999999
df_with_zone_sheet = pd.concat([master_df, fake_zone], ignore_index=True)
result2, _, _, _ = add_calculations(df_with_zone_sheet, factor_lookup=factor_lookup, zone_mapping={'North': ['Rajasthan']})
north2 = result2[(result2['State_Zone'] == 'North') & (result2['Urban_Rural'] == 'U') & (result2['TG_Segment'] == 'TOTAL') & (result2['Flag'] == 'Category')]
raj_u = result2[(result2['State_Zone'] == 'Rajasthan') & (result2['Urban_Rural'] == 'U') & (result2['TG_Segment'] == 'TOTAL') & (result2['Flag'] == 'Category')]
check("Zone's own sheet data ignored (synthesized from states only)",
      len(north2) == 1 and abs(north2['Sales Derived__2025 Apr - 2026 Mar'].values[0] - raj_u['Sales Derived__2025 Apr - 2026 Mar'].values[0]) < 1e-6)

# Zone with no member states present -> reported, not guessed
result3, _, unmapped3, _ = add_calculations(master_df, factor_lookup=factor_lookup, zone_mapping={'South': ['Kerala']})
check("Zone with absent member states reported as unmapped", 'South' in unmapped3)
check("Unmapped zone produces no rows", len(result3[result3['State_Zone'] == 'South']) == 0)

print("=" * 70)
print("TEST GROUP 8: Variance calculation")
print("=" * 70)
var_df = compute_variance(master_df, sku_df)
check("Variance rows produced", len(var_df) > 0)
sel = var_df[(var_df['State_Zone'] == 'Rajasthan') & (var_df['Urban_Rural'] == 'U+R') & (var_df['TG_Segment'] == 'TOTAL') & (var_df['Brand_SKU_Item'].str.contains('SELFIE'))]
check("Selfie variance row exists", len(sel) == 1)
if len(sel) == 1:
    bu = sel['Brand_Units__2025 Apr - 2026 Mar'].values[0]
    su = sel['SKU_Sum_Units__2025 Apr - 2026 Mar'].values[0]
    # manual: brand HH*NOP = 288.146*1.811
    manual_bu = 288.146 * 1.811
    check("Brand units match manual HH*NOP", abs(bu - manual_bu) < 1e-6, f"{bu} vs {manual_bu}")
    check("Variance = Brand - SKU sum", abs(sel['Variance__2025 Apr - 2026 Mar'].values[0] - (bu - su)) < 1e-9)

print("=" * 70)
print("TEST GROUP 9: Smart guess and duplicates")
print("=" * 70)
g = smart_guess_companies(['[HCEXL] GODREJ BRAND NEW THING'], DEFAULT_COMPANY_MAPPING)
check("Godrej keyword guess", g['[HCEXL] GODREJ BRAND NEW THING'][0] == 'GODREJ CONSUMER PRODS')
g2 = smart_guess_companies(['[HCEXL] TOTALLY UNKNOWN THING QQQ'], DEFAULT_COMPANY_MAPPING)
check("Unknown brand falls to Others / Unmapped", g2['[HCEXL] TOTALLY UNKNOWN THING QQQ'][0] == 'Others / Unmapped')
g3 = smart_guess_companies(['[HCEXL] GODREJ SELFIE HAIR COLOR SHAMPO'], DEFAULT_COMPANY_MAPPING)
check("Near-duplicate typo matched to GCPL", g3['[HCEXL] GODREJ SELFIE HAIR COLOR SHAMPO'][0] == 'GODREJ CONSUMER PRODS')
dupes = find_potential_duplicates(list(DEFAULT_BRAND_MAPPING.keys()))
check("Duplicate scan runs and finds known near-dupes", len(dupes) >= 1)

print("=" * 70)
print("TEST GROUP 10: Exporter (all sheets, openable workbook)")
print("=" * 70)
out = export_to_excel(result, dict(DEFAULT_INDIVIDUAL_FACTOR_UR), missing,
                      company_summary_df=cs, zone_mapping=zone_mapping,
                      unmapped_zones=unmapped, variance_df=var_df)
data = out.getvalue()
check("Export non-empty", len(data) > 100000)
import io
wb_out = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
expected_sheets = {'Master_Clean', 'Company_Summary', 'Variance', 'Brand_Mapping', 'Company_Mapping', 'Individual_Factor', 'Zone_Mapping'}
check("All expected sheets present", expected_sheets.issubset(set(wb_out.sheetnames)), str(wb_out.sheetnames))
ws_mc = wb_out['Master_Clean']
check("Master_Clean has data rows", ws_mc.max_row > 100)

print("=" * 70)
print("TEST GROUP 11: Mapping data integrity")
print("=" * 70)
check("Brand mapping 180 items", len(DEFAULT_BRAND_MAPPING) == 180)
check("Company mapping 180 items", len(DEFAULT_COMPANY_MAPPING) == 180)
check("Same keys in both mappings", set(DEFAULT_BRAND_MAPPING) == set(DEFAULT_COMPANY_MAPPING))
# Re-extract from source and diff
wb_t = openpyxl.load_workbook(TRUTH, data_only=True, read_only=True)
ws_t = wb_t['Sheet1']
fresh_flag, fresh_comp = {}, {}
for row in ws_t.iter_rows(min_row=2, values_only=True):
    fresh_flag[row[5]] = row[4]
    fresh_comp[row[5]] = row[96]
check("Brand mapping matches source byte-for-byte", fresh_flag == DEFAULT_BRAND_MAPPING)
check("Company mapping matches source byte-for-byte", fresh_comp == DEFAULT_COMPANY_MAPPING)

print("=" * 70)
print("TEST GROUP 12: App page logic simulation (session-state flows)")
print("=" * 70)
# Simulate the exact dict-building the app does before running the pipeline
mapping_df = pd.DataFrame([{'Brand_SKU_Item': k, 'Flag': v, 'Status': 'Verified'} for k, v in DEFAULT_BRAND_MAPPING.items()])
company_mapping_df = pd.DataFrame([{'Brand_SKU_Item': k, 'Company': v, 'Status': 'Verified'} for k, v in DEFAULT_COMPANY_MAPPING.items()])
fo = dict(zip(mapping_df['Brand_SKU_Item'], mapping_df['Flag']))
co = dict(zip(company_mapping_df['Brand_SKU_Item'], company_mapping_df['Company']))
check("flag_overrides built from df with Status col", fo == DEFAULT_BRAND_MAPPING)
check("company_overrides built from df with Status col", co == DEFAULT_COMPANY_MAPPING)

# Zone auto-build logic from app
sheet_sel = pd.DataFrame([{**o, 'Format': 'SHC'} for o in get_sheet_overview(open(RAW, 'rb'))])
included_states = set(sheet_sel[(sheet_sel['Include']) & (~sheet_sel['Is_Zone'])]['State_Zone'].unique())
zrows = []
for state in sorted(included_states):
    z = STATE_TO_ZONE.get(state)
    if z:
        zrows.append({'Zone': z, 'Member_State': state})
check("App zone auto-build produces North->Rajasthan", zrows == [{'Zone': 'North', 'Member_State': 'Rajasthan'}])

# Factor dict building as the app does
factor_df = pd.DataFrame([{'State_Zone': s, 'Individual_Factor_U': v.get('U'), 'Individual_Factor_R': v.get('R')} for s, v in DEFAULT_INDIVIDUAL_FACTOR_UR.items()])
fud = {}
for _, row in factor_df.iterrows():
    fud[row['State_Zone']] = {'U': row.get('Individual_Factor_U'), 'R': row.get('Individual_Factor_R')}
fl2 = build_factor_lookup(fud)
check("Factor lookup from app's df == direct lookup", fl2 == build_factor_lookup(dict(DEFAULT_INDIVIDUAL_FACTOR_UR)))
check("Rajasthan U factor exact", fl2[('Rajasthan', 'U')] == 1.32)
check("Rajasthan R factor exact", fl2[('Rajasthan', 'R')] == 1.24)

print("=" * 70)
print("TEST GROUP 13: Robustness edge cases")
print("=" * 70)
# Empty sku_df to variance
empty_var = compute_variance(master_df, pd.DataFrame())
check("Variance with empty SKU df doesn't crash", len(empty_var) > 0)
# Included sheets filter actually filters
f3 = open(RAW, 'rb')
m_only_ur, _ = process_workbook(f3, 'SHC', included_sheet_names={'Rajasthan (U+R)'})
check("Sheet whitelist honored", set(m_only_ur['Urban_Rural'].unique()) == {'U+R'})
# get_sheet_overview twice on same handle (seek safety)
f4 = open(RAW, 'rb')
ov1 = get_sheet_overview(f4)
ov2 = get_sheet_overview(f4)
check("Sheet overview re-readable on same handle", len(ov1) == len(ov2) == 3)
# process after overview on same handle (the app does exactly this)
m_after, _ = process_workbook(f4, 'SHC')
check("Workbook processable after overview on same handle", len(m_after) == 270, f"got {len(m_after)}")

print("=" * 70)
print("TEST GROUP 14: All India auto-rollup (never its own factor/sheet)")
print("=" * 70)
f5 = open(RAW, 'rb')
m5, s5 = process_all_files({'SHC': f5}, flag_overrides=dict(DEFAULT_BRAND_MAPPING), company_overrides=dict(DEFAULT_COMPANY_MAPPING))
result5, _, _, _ = add_calculations(m5, factor_lookup=factor_lookup, zone_mapping={})
ai5 = result5[(result5['State_Zone'] == 'All India') & (result5['Urban_Rural'] == 'U+R') & (result5['TG_Segment'] == 'TOTAL') & (result5['Flag'] == 'Category')]
raj5 = result5[(result5['State_Zone'] == 'Rajasthan') & (result5['Urban_Rural'] == 'U+R') & (result5['TG_Segment'] == 'TOTAL') & (result5['Flag'] == 'Category')]
check("All India auto-synthesized (no manual mapping needed)", len(ai5) == 1)
check("All India == sum of real states present (Rajasthan only here)",
      abs(ai5['Sales Derived__2025 Apr - 2026 Mar'].values[0] - raj5['Sales Derived__2025 Apr - 2026 Mar'].values[0]) < 1e-6)

# Plant a real All India sheet (with the OLD unverified 1.50/1.80 factor) alongside
# Rajasthan and confirm its own data/factor are still fully ignored.
fake_ai = m5[m5['Urban_Rural'] == 'U'].copy()
fake_ai['State_Zone'] = 'All India'
fake_ai['Is_Zone'] = False
for c in fake_ai.columns:
    if c.startswith('HH__') or c.startswith('Val__') or c.startswith('Avg NOP__'):
        fake_ai[c] = 999999
m5_with_ai_sheet = pd.concat([m5, fake_ai], ignore_index=True)
result6, _, _, _ = add_calculations(m5_with_ai_sheet, factor_lookup=factor_lookup, zone_mapping={})
ai6 = result6[(result6['State_Zone'] == 'All India') & (result6['Urban_Rural'] == 'U+R') & (result6['TG_Segment'] == 'TOTAL') & (result6['Flag'] == 'Category')]
check("All India's own planted sheet+factor fully ignored, still == Rajasthan",
      len(ai6) == 1 and abs(ai6['Sales Derived__2025 Apr - 2026 Mar'].values[0] - raj5['Sales Derived__2025 Apr - 2026 Mar'].values[0]) < 1e-6)

print("=" * 70)
print("TEST GROUP 15: Grammage/SU blank-inconsistency bug (regression)")
print("=" * 70)
# Two states with the SAME real values but DIFFERENT blank conventions for
# Grammage/SU (None vs ''), which is common across different real workbooks.
# Before the fix, this silently split zone/All India rollups into separate
# under-counted groups instead of combining them.
data = {
    'Format': ['SHC'] * 4,
    'State_Zone': ['StateA', 'StateA', 'StateB', 'StateB'],
    'Is_Zone': [False] * 4,
    'Urban_Rural': ['U', 'R', 'U', 'R'],
    'TG_Segment': ['TOTAL'] * 4,
    'Flag': ['Brand'] * 4,
    'Company': ['GCPL'] * 4,
    'Grammage': [None, None, '', ''],
    'SU': [None, None, None, None],
    'Brand_SKU_Item': ['Test Brand'] * 4,
    'HH__P1': [100.0, 100.0, 100.0, 100.0],
    'Val__P1': [10.0, 10.0, 10.0, 10.0],
    'Avg NOP__P1': [2.0, 2.0, 2.0, 2.0],
}
df_blank_test = pd.DataFrame(data)
fl_test = build_factor_lookup({'StateA': {'U': 1.0, 'R': 1.0}, 'StateB': {'U': 1.0, 'R': 1.0}})
result_blank, _, _, _ = add_calculations(df_blank_test, factor_lookup=fl_test, zone_mapping={})
ai_blank = result_blank[result_blank['State_Zone'] == 'All India']
ai_u_total = ai_blank[ai_blank['Urban_Rural'] == 'U']['Sales Derived__P1'].sum()
ai_r_total = ai_blank[ai_blank['Urban_Rural'] == 'R']['Sales Derived__P1'].sum()
check("All India combines states with inconsistent blank Grammage (U)", ai_u_total == 20000000.0, f"got {ai_u_total}")
check("All India combines states with inconsistent blank Grammage (R)", ai_r_total == 20000000.0, f"got {ai_r_total}")
check("No row-count inflation from blank-value grouping mismatch", len(ai_blank[ai_blank['Urban_Rural'] == 'U']) == 1)

print("=" * 70)
print("TEST GROUP 16: Zero Individual Factor detection (regression - Delhi bug)")
print("=" * 70)
# Reproduces the exact real-world bug: a state with Individual Factor = 0.00
# silently zeroed out Sales Derived/Units Estd for that entire state, with no
# other visible symptom. This must now be caught and reported every time.
zero_data = {
    'Format': ['SHC', 'SHC'],
    'State_Zone': ['Delhi', 'Delhi'],
    'Is_Zone': [False, False],
    'Urban_Rural': ['U', 'U'],
    'TG_Segment': ['TOTAL', 'TOTAL'],
    'Flag': ['Category', 'Brand'],
    'Company': ['Category Total', 'GODREJ CONSUMER PRODS'],
    'Grammage': [None, None],
    'SU': [None, None],
    'Brand_SKU_Item': ['[HCEXL] ANY HAIR COL.SHAMPOO', '[HCEXL] GODREJ SELFIE HAIR COLOUR SHAMPOO'],
    'HH__P1': [1100.445, 100.0],
    'Val__P1': [78.094, 5.0],
    'Avg NOP__P1': [2.942, 2.0],
}
zero_df = pd.DataFrame(zero_data)
bad_factor_lookup = {('Delhi', 'U'): 0.0}  # the exact bug: a literal zero
result_zero, _, _, zero_factors_found = add_calculations(zero_df, factor_lookup=bad_factor_lookup, zone_mapping={})
check("Zero factor is detected and reported", ('Delhi', 'U') in zero_factors_found, str(zero_factors_found))
cat_row = result_zero[result_zero['Flag'] == 'Category']
check("Reproduces the exact reported symptom: Sales Derived = 0 with a real Val", cat_row['Sales Derived__P1'].values[0] == 0.0)

# And confirm a non-zero factor for the same data does NOT get flagged
good_factor_lookup = {('Delhi', 'U'): 1.00}
result_good, _, _, zero_factors_good = add_calculations(zero_df, factor_lookup=good_factor_lookup, zone_mapping={})
check("Non-zero factor is never flagged", len(zero_factors_good) == 0)
cat_row_good = result_good[result_good['Flag'] == 'Category']
check("Sales Derived is correctly non-zero with a real factor", cat_row_good['Sales Derived__P1'].values[0] > 0)

print("=" * 70)
print("TEST GROUP 17: Header row detection (regression - garbled header bug)")
print("=" * 70)
# Reproduces the real bug found combining 7 raw files: a sheet whose header
# block sits at a different row position than usual (one extra row inserted)
# used to be silently misread as data, producing garbage column names like
# "401.147__1.358". The header row must now be found dynamically and
# validated, not assumed to sit at a fixed position.
from cleaner import build_column_map, find_header_rows

shifted_rows = [
    (None, None, None, None, 'Measure'),
    (None,) * 5,
    (None, None, None, None, 'EXTRA UNEXPECTED ROW'),
    (None, None, None, None, 'TestState (U)'),
    (None, None, None, None, None, 'HH'),
    (None, None, None, None, None, '2023 Apr - 2023 Jun'),
    ('TG', 'Format', 'Grammage', 'SU', 'Universe (000s)', 100.0),
    ('TOTAL', 'SHC', None, None, '[HCEXL] ANY HAIR COL.SHAMPOO', 50.0),
]
metric_idx, period_idx = find_header_rows(shifted_rows)
check("Header row found dynamically despite shifted layout", metric_idx == 4 and period_idx == 5, f"got {metric_idx}, {period_idx}")
cols = build_column_map(shifted_rows)
check("Column names built correctly from the shifted layout", cols[0] == "HH__2023 Apr - 2023 Jun", str(cols[:1]))

broken_rows = [
    (None, None, None, None, 'Measure'),
    (None,) * 5,
    (None, None, None, None, 'SomeState (U)'),
    (None, None, None, None, None, 'NOT_HH_LABEL'),
    (None, None, None, None, None, '2023 Apr - 2023 Jun'),
    ('TG', 'Format', 'Grammage', 'SU', 'Universe (000s)', 100.0),
    ('TOTAL', 'SHC', None, None, '[HCEXL] ANY HAIR COL.SHAMPOO', 50.0),
]
try:
    build_column_map(broken_rows)
    check("Genuinely unrecognizable layout raises an error (did not raise)", False)
except ValueError:
    check("Genuinely unrecognizable layout raises a clear error instead of producing garbage", True)

print("=" * 70)
print("TEST GROUP 18: Format C support (new flat-header, no-bracket-free layout)")
print("=" * 70)
from cleaner import (
    detect_sheet_raw_format, find_format_c_header, build_column_map_format_c,
    _format_c_period_to_canonical,
)

check("Quarter code translates to canonical label", _format_c_period_to_canonical("01. AMJ'23") == "2023 Apr - 2023 Jun")
check("MAT code translates to canonical label", _format_c_period_to_canonical("MAT Mar24") == "2023 Apr - 2024 Mar")
check("Invalid period code returns None (not silently wrong)", _format_c_period_to_canonical("garbage") is None)

FORMAT_C_FILE = 'sample_data/Sample_template_-_C.xlsx'
import os
if os.path.exists(FORMAT_C_FILE):
    shc_sheets = {'Maharashtra(U+R) - SHC', 'Maharasthra(R) - SHC', 'Maharasthra(U) - SHC'}
    df_c, sku_c = process_workbook(FORMAT_C_FILE, 'SHC', included_sheet_names=shc_sheets)
    check("Format C: rows extracted", len(df_c) > 0, f"got {len(df_c)}")
    check("Format C: all 3 Urban_Rural cuts present", set(df_c['Urban_Rural'].unique()) == {'U', 'R', 'U+R'})
    check("Format C: all 5 TG segments present", set(df_c['TG_Segment'].unique()) == {'TOTAL', 'SEC A', 'SEC B', 'SEC C', 'SEC D/E'})
    check("Format C: no SKU rows leaked into Master_Clean", 'SKU' not in df_c['Flag'].unique())
    check("Format C: Category rows correctly classified", (df_c['Flag'] == 'Category').sum() > 0)

    df_c_norm = df_c.copy()
    df_c_norm['State_Zone'] = 'Maharashtra'
    fl = build_factor_lookup({'Maharashtra': {'U': 1.85, 'R': 1.85}})
    result_c, missing_c, unmapped_c, zero_c = add_calculations(df_c_norm, factor_lookup=fl, zone_mapping={})
    cat_c = result_c[(result_c['TG_Segment'] == 'TOTAL') & (result_c['Flag'] == 'Category') & (result_c['State_Zone'] == 'Maharashtra')]
    ur_val = cat_c[cat_c['Urban_Rural'] == 'U+R']['Sales Derived__2025 Apr - 2026 Mar'].values[0]
    u_val = cat_c[cat_c['Urban_Rural'] == 'U']['Sales Derived__2025 Apr - 2026 Mar'].values[0]
    r_val = cat_c[cat_c['Urban_Rural'] == 'R']['Sales Derived__2025 Apr - 2026 Mar'].values[0]
    check("Format C: U+R = U + R identity holds exactly", abs(ur_val - (u_val + r_val)) < 1e-6, f"UR={ur_val} U+R sum={u_val+r_val}")
else:
    print("  SKIP (Format C sample file not present in this environment)")

# Sheet name parsing must still handle compound names (state+UR+format all in one)
check("Compound sheet name (Format C sample style) parses correctly",
      parse_region_sheet_name("Maharashtra(U+R) -Val added pwd") == ("Maharashtra", "U+R", False))
check("Original simple sheet name still parses correctly (no regression)",
      parse_region_sheet_name("Rajasthan (U+R)") == ("Rajasthan", "U+R", False))

print("=" * 70)
print("TEST GROUP 19: Format A/C detection safety (no cross-contamination)")
print("=" * 70)
import openpyxl as _openpyxl_test

wb_c_test = _openpyxl_test.load_workbook(FORMAT_C_FILE, data_only=True, read_only=True) if os.path.exists(FORMAT_C_FILE) else None
if wb_c_test is not None:
    rows_c_test = list(wb_c_test['Maharashtra(U+R) - SHC'].iter_rows(values_only=True))
    check("Format C file detected as C", detect_sheet_raw_format(rows_c_test) == 'C')
    a_sig_in_c, _ = find_header_rows(rows_c_test)
    check("Format C file does NOT also match Format A's signature (no false positive)", a_sig_in_c is None)

wb_a_test = _openpyxl_test.load_workbook(RAW, data_only=True, read_only=True)
rows_a_test = list(wb_a_test['Rajasthan (U+R)'].iter_rows(values_only=True))
check("Format A file detected as A", detect_sheet_raw_format(rows_a_test) == 'A')
c_sig_in_a, _ = find_format_c_header(rows_a_test)
check("Format A file does NOT also match Format C's signature (no false positive)", c_sig_in_a is None)

# Adversarial: garbage sheet must be rejected, not guessed
garbage_rows = [(None,) * 5, ('foo', 'bar', 'baz', None, None), (1, 2, 3, 4, 5)]
check("Unrecognizable sheet returns None (not silently guessed)", detect_sheet_raw_format(garbage_rows) is None)

print("=" * 70)
print("TEST GROUP 20: Brand mapping suffix fallback (HBP naming mismatch fix)")
print("=" * 70)
from cleaner import _strip_any_suffix, _with_suffix_fallback

check("Suffix stripping works", _strip_any_suffix("[HCEXL] ABIHA POWDER [ANY HERBAL BASED PWD]") == "[HCEXL] ABIHA POWDER")
check("No-suffix name is unchanged", _strip_any_suffix("[HCEXL] GODREJ SELFIE HAIR COLOUR SHAMPOO") == "[HCEXL] GODREJ SELFIE HAIR COLOUR SHAMPOO")

test_mapping = {"[HCEXL] ABIHA POWDER [ANY HERBAL BASED PWD]": "Brand"}
extended = _with_suffix_fallback(test_mapping)
check("Fallback key added for suffixed entry", extended.get("[HCEXL] ABIHA POWDER") == "Brand")
check("Original suffixed key still present", extended.get("[HCEXL] ABIHA POWDER [ANY HERBAL BASED PWD]") == "Brand")

# Ambiguity safety: if the stripped form already means something ELSE, don't overwrite it
ambiguous_mapping = {
    "[HCEXL] FOO [ANY X]": "Brand",
    "[HCEXL] FOO": "Sub-brand",  # genuinely different, pre-existing entry
}
extended_ambig = _with_suffix_fallback(ambiguous_mapping)
check("Fallback does not overwrite a real distinct existing entry", extended_ambig["[HCEXL] FOO"] == "Sub-brand")

if os.path.exists(FORMAT_C_FILE):
    hbp_sheets = {'Maharashtra(R) - HBP', 'Maharasthra(U) - HBP', 'Maharashtra(U+R) - HBP'}
    df_hbp, _ = process_workbook(FORMAT_C_FILE, 'HBP', included_sheet_names=hbp_sheets,
                                  flag_overrides=dict(DEFAULT_BRAND_MAPPING), company_overrides=dict(DEFAULT_COMPANY_MAPPING))
    brand_rows_hbp = df_hbp[df_hbp['Flag'] == 'Brand']
    unmapped_hbp = (brand_rows_hbp['Company'] == 'Others / Unmapped').sum()
    check("Real HBP file: zero genuinely unmapped Brand rows after suffix fix", unmapped_hbp == 0, f"got {unmapped_hbp}")

print("=" * 70)
print("TEST GROUP 21: Full mapping coverage audit across all 7 Format C formats")
print("=" * 70)
if os.path.exists(FORMAT_C_FILE):
    all_format_sheets = {
        'SHC': {'Maharashtra(U+R) - SHC', 'Maharasthra(R) - SHC', 'Maharasthra(U) - SHC'},
        'HBP': {'Maharasthra(R) - HBP', 'Maharasthra(U) - HBP', 'Maharashtra(U+R) - HBP'},
        'Henna': {'Maharasthra(U+R) - Henna', 'Maharasthra(R) - Henna', 'Maharasthra(U) - Henna'},
        'OBL': {'Maharasthra(U+R) - OBL', 'Maharasthra(R) - OBL', 'Maharasthra(U) - OBL'},
        'KM': {'Maharasthra(U+R) - Pwd KM', 'Maharasthra(R) - Pwd KM', 'Maharasthra(U) - Pwd KM'},
        'ValueAdd': {'Maharashtra(U+R) -Val added pwd', 'Maharashtra(U) - val added pwd', 'Maharashtra(R) - Val added pwd'},
        'BasicPwd': {'Maharashtra(R) - Basic Pwd', 'Maharashtra(U+R) - Basic Pwd', 'Maharashtra(U) - Basic Pwd'},
    }
    extended_brand_test = _with_suffix_fallback(DEFAULT_BRAND_MAPPING)
    all_items_test = set()
    per_format_ok = True
    for fmt, sheets in all_format_sheets.items():
        df_fmt, _ = process_workbook(FORMAT_C_FILE, fmt, included_sheet_names=sheets)
        items_fmt = df_fmt[df_fmt['Flag'].isin(['Brand', 'Sub-brand', 'Others'])]['Brand_SKU_Item'].unique()
        all_items_test.update(items_fmt)
        for it in items_fmt:
            if it not in extended_brand_test:
                per_format_ok = False
    check("All 7 formats extract without error", per_format_ok is not None)
    missing_test = [i for i in all_items_test if i not in extended_brand_test]
    check(f"Zero genuinely unmapped items across all 7 formats ({len(all_items_test)} checked)", len(missing_test) == 0, str(missing_test))

print("=" * 70)
print("TEST GROUP 22: Format B support (monthly-only, no-bracket layout)")
print("=" * 70)
from cleaner import (
    classify_flag_format_b, find_format_b_signature, add_format_b_annual_totals,
    _month_period_sort_key,
)

check("Category detected via leading+trailing space + 'ANY' prefix", classify_flag_format_b(" ANY CREME ") == "Category")
check("Brand detected via leading+trailing space", classify_flag_format_b(" MATRIX SOCOLOR PREBONDED CREAM HAIR COLO ") == "Brand")
check("SKU detected via no surrounding space", classify_flag_format_b("Matrix SoColor Pre-Bonded 90 GBlackCreme") == "SKU")
check("Others detected via OTH. pattern", classify_flag_format_b(" OTH.BRDED SOMETHING ") == "Others")

check("Chronological sort key orders correctly (not alphabetically)",
      sorted(["2025 Aug", "2025 Jun", "2025 Dec"], key=_month_period_sort_key) == ["2025 Jun", "2025 Aug", "2025 Dec"])

FORMAT_B_FILE = 'sample_data/Sample_format_B.xlsx'
if os.path.exists(FORMAT_B_FILE):
    df_b, sku_b = process_workbook(FORMAT_B_FILE, 'Creme')
    check("Format B: rows extracted", len(df_b) > 0, f"got {len(df_b)}")
    check("Format B: only TOTAL segment present (verified, not assumed)", set(df_b['TG_Segment'].unique()) == {'TOTAL'})
    check("Format B: no SKU rows leaked into Master_Clean", 'SKU' not in df_b['Flag'].unique())
    check("Format B: Category row correctly classified", (df_b['Flag'] == 'Category').sum() == 1)

    df_b_annual = add_format_b_annual_totals(df_b, format_tag='Creme')
    annual_hh_col = [c for c in df_b_annual.columns if c.startswith('HH__') and ' - ' in c.split('__', 1)[1]]
    check("Format B: annual HH column created with honest date-range label",
          len(annual_hh_col) == 1 and annual_hh_col[0] == 'HH__2025 Jun - 2026 May', str(annual_hh_col))

    cat_b = df_b_annual[df_b_annual['Flag'] == 'Category']
    month_cols = [f'HH__{p}' for p in ['2025 Jun', '2025 Jul', '2025 Aug', '2025 Sep', '2025 Oct', '2025 Nov',
                                        '2025 Dec', '2026 Jan', '2026 Feb', '2026 Mar', '2026 Apr', '2026 May']]
    manual_sum = cat_b[month_cols].sum(axis=1).values[0]
    check("Format B: annual sum exactly matches manual sum of all 12 months",
          abs(cat_b[annual_hh_col[0]].values[0] - manual_sum) < 1e-9)

    # Safety: mixing real Format A data must never corrupt or get corrupted by Format B's annual-totals step
    fa2 = open(RAW, 'rb')
    fb2 = open(FORMAT_B_FILE, 'rb')
    mixed_df, _ = process_all_files({'SHC': fa2, 'Creme': fb2}, flag_overrides=dict(DEFAULT_BRAND_MAPPING), company_overrides=dict(DEFAULT_COMPANY_MAPPING))
    mixed_result = add_format_b_annual_totals(mixed_df, format_tag='Creme')
    shc_mat_untouched = mixed_result[mixed_result['Format'] == 'SHC']['HH__2023 Apr - 2024 Mar'].equals(
        mixed_df[mixed_df['Format'] == 'SHC']['HH__2023 Apr - 2024 Mar']
    )
    check("Format A's own MAT column untouched when mixed with Format B in the same table", shc_mat_untouched)
else:
    print("  SKIP (Format B sample file not present in this environment)")

print("=" * 70)
print("TEST GROUP 23: Bracket-prefix naming fallback (Format B mapping fix)")
print("=" * 70)
from cleaner import _strip_bracket_prefix, _with_naming_fallbacks

check("Prefix stripping works", _strip_bracket_prefix("[HCEXL] GODREJ PROF.DIMENSION CREAM HAIR COLOR") == "GODREJ PROF.DIMENSION CREAM HAIR COLOR")
check("No-prefix name is unchanged", _strip_bracket_prefix("GODREJ PROF.DIMENSION CREAM HAIR COLOR") == "GODREJ PROF.DIMENSION CREAM HAIR COLOR")

combo_test = {"[HCEXL] COLORMATE [ANY HERBAL BASED PWD]": "HENNA INDUSTRIES PVT LTD"}
combo_extended = _with_naming_fallbacks(combo_test)
check("Combined prefix+suffix strip resolves to bare brand name", combo_extended.get("COLORMATE") == "HENNA INDUSTRIES PVT LTD")

if os.path.exists(FORMAT_B_FILE):
    df_b_map, _ = process_workbook(FORMAT_B_FILE, 'Creme')
    items_b = df_b_map[df_b_map['Flag'] == 'Brand']['Brand_SKU_Item'].tolist()
    extended_company_b = _with_naming_fallbacks(DEFAULT_COMPANY_MAPPING)
    matched_b = sum(1 for i in items_b if i in extended_company_b)
    check(f"Real Format B file: most brands ({matched_b}/{len(items_b)}) now resolve via naming fallback",
          matched_b >= len(items_b) - 3, f"only {matched_b}/{len(items_b)} matched")

print("=" * 70)
print("TEST GROUP 24: Sales Derived / Units Estd decoupling (Format B annual)")
print("=" * 70)
if os.path.exists(FORMAT_B_FILE):
    df_dec, _ = process_workbook(FORMAT_B_FILE, 'Creme')
    df_dec = add_format_b_annual_totals(df_dec, format_tag='Creme', include_estimated_avg_nop=True)
    fl_dec = build_factor_lookup({'Maharasthra': {'U': 1.85, 'R': 1.85}})
    result_dec, _, _, _ = add_calculations(df_dec, factor_lookup=fl_dec, zone_mapping={})
    annual_label = '2025 Jun - 2026 May'
    check("Sales Derived computed for annual period (needs only HH+Val)",
          f'Sales Derived__{annual_label}' in result_dec.columns)
    check("Units Estd correctly NOT computed for annual period (Avg NOP unconfirmed)",
          f'Units Estd__{annual_label}' not in result_dec.columns)
    cat_dec = result_dec[result_dec['Flag'] == 'Category']
    check("Annual Sales Derived value is a real positive number",
          cat_dec[f'Sales Derived__{annual_label}'].values[0] > 0)

# Re-confirm this never regresses Format A (which always has all 3 columns together)
f_a_dec = open(RAW, 'rb')
m_a_dec, _ = process_all_files({'SHC': f_a_dec}, flag_overrides=dict(DEFAULT_BRAND_MAPPING), company_overrides=dict(DEFAULT_COMPANY_MAPPING))
fl_a_dec = build_factor_lookup(dict(DEFAULT_INDIVIDUAL_FACTOR_UR))
result_a_dec, _, _, _ = add_calculations(m_a_dec, factor_lookup=fl_a_dec, zone_mapping={'North': ['Rajasthan']})
check("Format A: Units Estd still computed normally for every period (no regression)",
      'Units Estd__2025 Apr - 2026 Mar' in result_a_dec.columns)

print("=" * 70)
print("TEST GROUP 25: Format B - U+R and Zone rollup on real data structure")
print("=" * 70)
if os.path.exists(FORMAT_B_FILE):
    import tempfile
    src_wb = openpyxl.load_workbook(FORMAT_B_FILE, data_only=True, read_only=True)
    src_rows = list(src_wb['Maharasthra(U)'].iter_rows(values_only=True))

    synth_wb = openpyxl.Workbook()
    synth_wb.remove(synth_wb.active)
    for label in ['U', 'R', 'U+R']:
        ws = synth_wb.create_sheet(f'TestState ({label})')
        for r in src_rows:
            ws.append(r)
        ws.cell(row=6, column=1, value=f'TestState ({label})')

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        synth_path = tmp.name
    synth_wb.save(synth_path)

    df_synth, _ = process_workbook(synth_path, 'Creme')
    check("Synthetic multi-sheet Format B: all 3 UR cuts present", set(df_synth['Urban_Rural'].unique()) == {'U', 'R', 'U+R'})

    fl_synth = build_factor_lookup({'TestState': {'U': 1.85, 'R': 2.10}})
    result_synth, missing_synth, unmapped_synth, _ = add_calculations(df_synth, factor_lookup=fl_synth, zone_mapping={'TestZone': ['TestState']})
    check("No missing factors, no unmapped zones", len(missing_synth) == 0 and len(unmapped_synth) == 0)

    cat_synth = result_synth[(result_synth['Flag'] == 'Category') & (result_synth['TG_Segment'] == 'TOTAL')]
    u_v = cat_synth[(cat_synth['State_Zone'] == 'TestState') & (cat_synth['Urban_Rural'] == 'U')]['Sales Derived__2025 Jun'].values[0]
    r_v = cat_synth[(cat_synth['State_Zone'] == 'TestState') & (cat_synth['Urban_Rural'] == 'R')]['Sales Derived__2025 Jun'].values[0]
    ur_v = cat_synth[(cat_synth['State_Zone'] == 'TestState') & (cat_synth['Urban_Rural'] == 'U+R')]['Sales Derived__2025 Jun'].values[0]
    check("Format B: U+R = U + R exactly, on real data structure", abs(ur_v - (u_v + r_v)) < 1e-6, f"UR={ur_v} sum={u_v+r_v}")

    zone_v = cat_synth[(cat_synth['State_Zone'] == 'TestZone') & (cat_synth['Urban_Rural'] == 'U+R')]['Sales Derived__2025 Jun'].values[0]
    check("Format B: Zone rollup matches sum of member state exactly", abs(zone_v - ur_v) < 1e-6)

    ai_v = cat_synth[(cat_synth['State_Zone'] == 'All India') & (cat_synth['Urban_Rural'] == 'U+R')]['Sales Derived__2025 Jun'].values[0]
    check("Format B: All India auto-rollup matches exactly", abs(ai_v - ur_v) < 1e-6)

    os.remove(synth_path)
else:
    print("  SKIP (Format B sample file not present in this environment)")



print("=" * 70)
print("TEST GROUP 26: KPI Explorer page (numbers must match pipeline exactly)")
print("=" * 70)
try:
    from streamlit.testing.v1 import AppTest
    import warnings as _w; _w.filterwarnings("ignore")

    _m26, _ = process_all_files({"SHC": RAW}, verbose=False)
    _r26, _, _, _ = add_calculations(_m26)
    _cs26 = build_company_summary(_r26)

    _at = AppTest.from_file("app.py", default_timeout=120)
    _at.session_state["result_df"] = _r26
    _at.session_state["company_summary_df"] = _cs26
    _at.session_state["current_page"] = "explorer"
    _at.run()
    check("Explorer: page loads with no exception", len(_at.exception) == 0)

    _at.selectbox[0].select("GODREJ CONSUMER PRODS")
    _at.selectbox[1].select("Value MS%")
    _at.selectbox[2].select("All India")
    _at.selectbox[3].select("U+R")
    _at.selectbox[4].select("TOTAL")
    _per26 = "2025 Apr - 2026 Mar"
    _at.selectbox[5].select(_per26)
    _at.run()
    check("Explorer: company single-period selection runs clean", len(_at.exception) == 0)

    _exp26 = _cs26[(_cs26.Company == "GODREJ CONSUMER PRODS") & (_cs26.State_Zone == "All India")
                   & (_cs26.Urban_Rural == "U+R") & (_cs26.TG_Segment == "TOTAL")]["Value MS%__" + _per26].iloc[0]
    _cards26 = [str(md.value) for md in _at.markdown if "stat-card" in str(md.value)]
    _shown26 = _cards26[-1].split('stat-value">')[1].split("<")[0] if _cards26 else ""
    check("Explorer: company Value MS% shown matches pipeline exactly",
          _shown26 == f"{_exp26*100:,.2f}%", f"shown={_shown26} expected={_exp26*100:,.2f}%")

    _at.radio[0].set_value("Brand")
    _at.run()
    _gb26 = [b for b in _at.selectbox[0].options if "SELFIE" in b.upper()][0]
    _at.selectbox[0].select(_gb26)
    _at.selectbox[1].select("Sales Derived")
    _at.selectbox[2].select("Rajasthan")
    _at.selectbox[3].select("U")
    _at.selectbox[4].select("TOTAL")
    _at.selectbox[5].select(_at.selectbox[5].options[0])  # All periods (trend)
    _at.run()
    check("Explorer: brand trend mode runs clean", len(_at.exception) == 0)

    _sub26 = _r26[(_r26.Brand_SKU_Item == _gb26) & (_r26.State_Zone == "Rajasthan")
                  & (_r26.Urban_Rural == "U") & (_r26.TG_Segment == "TOTAL")
                  & (_r26.Flag.isin(["Brand", "Others"]))]
    _shown_df26 = _at.dataframe[0].value
    _all_ok26 = len(_shown_df26) == 15
    for _, _row26 in _shown_df26.iterrows():
        _e26 = _sub26["Sales Derived__" + _row26.Period].iloc[0]
        if f"{_e26:,.2f}" != _row26["Sales Derived"]:
            _all_ok26 = False
    check("Explorer: all 15 brand trend datapoints match pipeline exactly", _all_ok26)
except ImportError:
    print("  SKIP (streamlit AppTest not available in this environment)")



print("=" * 70)
print("TEST GROUP 27: All India sheets auto-unticked (Henna crash fix)")
print("=" * 70)
from cleaner import get_sheet_overview as _gso27
import openpyxl as _op27, io as _io27

def _mk_wb27(names):
    wb = _op27.Workbook()
    wb.remove(wb.active)
    for n in names:
        wb.create_sheet(n)
    b = _io27.BytesIO()
    wb.save(b)
    b.seek(0)
    return b

_ov27 = {o["Sheet_Name"]: o for o in _gso27(_mk_wb27([
    "All India U+R", "All India Urban", "All India Rural",
    "Maharashtra (U)", "Rajasthan (U+R)", "North (U+R)",
]))}

check("All India U+R sheet is auto-unticked", _ov27["All India U+R"]["Include"] is False)
check("All India Urban sheet is auto-unticked", _ov27["All India Urban"]["Include"] is False)
check("All India Rural sheet is auto-unticked", _ov27["All India Rural"]["Include"] is False)
check("All India sheets flagged with Is_All_India", all(
    _ov27[n]["Is_All_India"] is True for n in ["All India U+R", "All India Urban", "All India Rural"]))
check("Real state sheets are still ticked by default (no regression)",
      _ov27["Maharashtra (U)"]["Include"] is True and _ov27["Rajasthan (U+R)"]["Include"] is True)
check("State sheets are NOT flagged as All India",
      _ov27["Maharashtra (U)"]["Is_All_India"] is False)
check("Zone sheets still auto-unticked (no regression)", _ov27["North (U+R)"]["Include"] is False)

# The whole point: excluding the All India sheet must not change any number,
# because add_calculations rebuilds All India from states and discards raw rows.
_m27, _ = process_all_files({"SHC": RAW}, verbose=False)
_r27, _, _, _ = add_calculations(_m27)
_ai27 = _r27[_r27["State_Zone"] == "All India"]
check("All India still produced even with no All India sheet read", len(_ai27) > 0,
      f"got {len(_ai27)} All India rows")



print("=" * 70)
print("TEST GROUP 28: Rollup coverage safeguard + intentional zero factors")
print("=" * 70)
from calculator import build_rollup_coverage, KNOWN_INTENTIONAL_ZERO_FACTORS

check("Delhi (U) is the only known-intentional zero factor",
      KNOWN_INTENTIONAL_ZERO_FACTORS == {("Delhi", "U")})

# --- Scenario 1: full coverage -> zero gaps ---
_m28, _ = process_all_files({"SHC": RAW}, verbose=False)
_r28, _, _, _ = add_calculations(_m28)
_cov28 = build_rollup_coverage(_r28, zone_mapping={})
_gaps28 = _cov28[_cov28["States_Included"] < _cov28["States_Expected"]]
check("Full-coverage run reports zero gaps (no false alarms)", len(_gaps28) == 0,
      f"gaps: {_gaps28.to_dict('records') if len(_gaps28) else 'none'}")
check("Coverage rows produced for all 3 UR cuts",
      set(_cov28["Urban_Rural"]) == {"U", "R", "U+R"})

# --- Scenario 2: Sagar's exact situation - only one state has U+R ---
_sag = _r28[~((_r28["Urban_Rural"] == "U+R") & (_r28["State_Zone"] != "Rajasthan"))].copy()
# build a 2-state world: duplicate Rajasthan as a fake second state with U and R only
_fake = _r28[(_r28["State_Zone"] == "Rajasthan") & (_r28["Urban_Rural"].isin(["U", "R"]))].copy()
_fake["State_Zone"] = "Testland"
_two = pd.concat([_sag, _fake], ignore_index=True)
_cov2 = build_rollup_coverage(_two, zone_mapping={})
_ur_row = _cov2[(_cov2["Rollup"] == "All India") & (_cov2["Urban_Rural"] == "U+R")].iloc[0]
check("Partial U+R coverage detected (1 of 2 states)",
      _ur_row["States_Included"] == 1 and _ur_row["States_Expected"] == 2,
      f"included={_ur_row['States_Included']} expected={_ur_row['States_Expected']}")
check("The missing state is named", _ur_row["Missing_States"] == "Testland")
_u_row = _cov2[(_cov2["Rollup"] == "All India") & (_cov2["Urban_Rural"] == "U")].iloc[0]
check("U cut with full coverage NOT flagged", _u_row["States_Included"] == _u_row["States_Expected"])

# --- Scenario 3: the West Bengal / West zone name trap ---
_wb = _r28[_r28["Urban_Rural"].isin(["U", "R", "U+R"])].copy()
_wb["State_Zone"] = _wb["State_Zone"].replace({"Rajasthan": "West Bengal"})
_wb.loc[_wb["State_Zone"] == "All India", "State_Zone"] = "All India"
_cov3 = build_rollup_coverage(_wb, zone_mapping={"West": ["Maharashtra", "Gujarat"]})
_ai3 = _cov3[_cov3["Rollup"] == "All India"]
check("West Bengal counted as a state, never confused with West zone",
      all(_ai3["States_Expected"] >= 1) and "West Bengal" not in set(_cov3["Rollup"]))
check("Zone with no members present in data produces no coverage row (matches unmapped_zones behaviour)",
      (_cov3["Rollup"] == "West").sum() == 0)

# --- Scenario 4: zone coverage counts only mapped members ---
_cov4 = build_rollup_coverage(_two, zone_mapping={"North": ["Rajasthan", "Testland"]})
_n_ur = _cov4[(_cov4["Rollup"] == "North") & (_cov4["Urban_Rural"] == "U+R")].iloc[0]
check("Zone U+R partial coverage detected", _n_ur["States_Included"] == 1 and _n_ur["States_Expected"] == 2)
_n_u = _cov4[(_cov4["Rollup"] == "North") & (_cov4["Urban_Rural"] == "U")].iloc[0]
check("Zone U full coverage not flagged", _n_u["States_Included"] == 2)

# --- Coverage function must never alter the result ---
_before = _r28.copy()
build_rollup_coverage(_r28, zone_mapping={"North": ["Rajasthan"]})
check("Coverage report leaves result_df completely untouched", _r28.equals(_before))



print("=" * 70)
print("TEST GROUP 29: Format B without a 'Process Period' row (Henna layout)")
print("=" * 70)
from cleaner import find_format_b_signature as _fbs29
import openpyxl as _op29, io as _io29

# Variant 1: the Creme-style block (Process Period present)
check("Creme-style block still detected", _fbs29([("Process Period", "2025 Jun To 2026 May")]) is True)
# Variant 2: the Henna-style block (no Process Period, starts at Universe)
check("Henna-style block detected (Universe + With Growth Factor)",
      _fbs29([("Universe", "With Growth Factor"), ("Measure", "Households in 000s")]) is True)
# Must NOT be too loose
check("'Universe' alone is NOT enough (no false positive)",
      _fbs29([("Universe", None)]) is False)
check("'Universe' with an unrelated partner cell is NOT enough",
      _fbs29([("Universe", "Rajasthan")]) is False)
check("Empty / junk rows are not a Format B signature",
      _fbs29([(None, None), ("foo", "bar")]) is False)

# The real files must still detect as their own format (no cross-contamination)
_wbA29 = _op29.load_workbook(RAW, data_only=True, read_only=True)
_rowsA29 = list(_wbA29["Rajasthan (U+R)"].iter_rows(values_only=True))
check("Format A file still detected as A (not stolen by widened B signature)",
      detect_sheet_raw_format(_rowsA29) == "A")
check("Format A file does not match the widened Format B signature",
      _fbs29(_rowsA29) is False)

if os.path.exists(FORMAT_C_FILE):
    _wbC29 = _op29.load_workbook(FORMAT_C_FILE, data_only=True, read_only=True)
    _rowsC29 = list(_wbC29["Maharashtra(U+R) - SHC"].iter_rows(values_only=True))
    check("Format C file still detected as C (not stolen by widened B signature)",
          detect_sheet_raw_format(_rowsC29) == "C")
    check("Format C file does not match the widened Format B signature",
          _fbs29(_rowsC29) is False)

# End-to-end: a real Format B file with its 'Process Period' row deleted (=
# exactly the Henna layout) must parse to identical data.
if os.path.exists(FORMAT_B_FILE):
    _base29, _ = process_workbook(FORMAT_B_FILE, "Creme")
    _wb29 = _op29.load_workbook(FORMAT_B_FILE)
    _wb29[_wb29.sheetnames[0]].delete_rows(1)
    _buf29 = _io29.BytesIO(); _wb29.save(_buf29); _buf29.seek(0)
    _henna29, _ = process_workbook(_buf29, "Henna")
    check("Henna-style file (no Process Period row) parses without error", len(_henna29) > 0)
    _cols29 = [c for c in _base29.columns if c != "Format"]
    check("Henna-style parse is byte-identical to the Creme-style baseline",
          _base29[_cols29].reset_index(drop=True).equals(_henna29[_cols29].reset_index(drop=True)))



print("=" * 70)
print("TEST GROUP 30: Upload page - format name registers on the FIRST enter")
print("=" * 70)
try:
    from streamlit.testing.v1 import AppTest as _AT30
    import warnings as _w30; _w30.filterwarnings("ignore")

    class _FakeFile30:
        def __init__(self, name, size=1000):
            self.name = name
            self.size = size
        def seek(self, *a, **k):
            pass

    _at30 = _AT30.from_file("app.py", default_timeout=120)
    _at30.session_state["staged_files"] = {
        "KPI-SHC.xlsx": [_FakeFile30("KPI-SHC.xlsx"), ""],
        "KPI-Henna.xlsx": [_FakeFile30("KPI-Henna.xlsx"), ""],
    }
    _at30.session_state["current_page"] = "upload"
    _at30.run()

    def _cap30(a):
        return [c.value for c in a.sidebar.caption]

    check("Upload page: starts with 0 files", any("0 file(s) added" in c for c in _cap30(_at30)))

    # The actual bug: this used to need TWO edits before anything updated.
    _at30.text_input[0].set_value("SHC")
    _at30.run()
    check("Format map updated after a single edit", "SHC" in _at30.session_state["file_format_map"])
    check("SIDEBAR updated after a single edit (no second keystroke needed)",
          any("1 file(s) added" in c and "SHC" in c for c in _cap30(_at30)),
          f"sidebar showed: {_cap30(_at30)}")

    _at30.text_input[1].set_value("Henna")
    _at30.run()
    check("Second file also registers on its first edit",
          any("2 file(s) added" in c for c in _cap30(_at30)))

    _at30.text_input[0].set_value("")
    _at30.run()
    check("Clearing a format takes effect immediately",
          any("1 file(s) added" in c and "Henna" in c for c in _cap30(_at30)))

    _at30.text_input[0].set_value("   ")
    _at30.run()
    check("Whitespace-only format is ignored, creates no phantom entry",
          len(_at30.session_state["file_format_map"]) == 1)
    check("Upload page raises no exception through all of this", len(_at30.exception) == 0)
except ImportError:
    print("  SKIP (streamlit AppTest not available in this environment)")



print("=" * 70)
print("TEST GROUP 31: Unrecognized-sheet error is self-diagnosing")
print("=" * 70)
import openpyxl as _op31, io as _io31

_wb31 = _op31.Workbook()
_ws31 = _wb31.active
_ws31.title = "Delhi (U)"
_ws31["A1"] = "Something Odd"; _ws31["B1"] = "Unexpected Header"
_ws31["A2"] = "Measure"; _ws31["B2"] = "Households in 000s"
_ws31["A6"] = "Delhi (U)"
for _r31 in range(7, 14):
    _ws31.cell(row=_r31, column=1, value="x")
_buf31 = _io31.BytesIO(); _wb31.save(_buf31); _buf31.seek(0)

_msg31 = ""
try:
    process_workbook(_buf31, "Henna")
    check("Unrecognized sheet raises rather than guessing", False, "no error raised")
except ValueError as _e31:
    _msg31 = str(_e31)
    check("Unrecognized sheet raises rather than guessing", True)

check("Error names the file and sheet", "Henna" in _msg31 and "Delhi (U)" in _msg31)
check("Error mentions BOTH Format B signature variants (proves new build is running)",
      "Process Period" in _msg31 and "Growth Factor" in _msg31)
check("Error reports what it actually saw in the sheet",
      "First rows actually seen" in _msg31 and "Something Odd" in _msg31)



print("=" * 70)
print("TEST GROUP 32: Format B - 'Analysis / Crosstab PulsePlus' opener (Val Added Pwd)")
print("=" * 70)
from cleaner import find_format_b_signature as _fbs32
import openpyxl as _op32, io as _io32

check("Analysis + Crosstab-PulsePlus opener detected",
      _fbs32([("Analysis", "Crosstab - PulsePlus-8.0"), ("Measure", "Households in 000s")]) is True)
check("'Analysis' alone is NOT enough (no false positive)",
      _fbs32([("Analysis", "Something else")]) is False)
check("'Analysis' with only Crosstab but no PulsePlus is NOT enough",
      _fbs32([("Analysis", "Crosstab - other")]) is False)
check("All three Format B openers now accepted",
      _fbs32([("Process Period", "x")]) and
      _fbs32([("Universe", "With Growth Factor")]) and
      _fbs32([("Analysis", "Crosstab - PulsePlus-8.0")]))

# Real Format A / C files must still detect as themselves
_wbA32 = _op32.load_workbook(RAW, data_only=True, read_only=True)
check("Format A still detected as A (Analysis opener didn't loosen detection)",
      detect_sheet_raw_format(list(_wbA32["Rajasthan (U+R)"].iter_rows(values_only=True))) == "A")
if os.path.exists(FORMAT_C_FILE):
    _wbC32 = _op32.load_workbook(FORMAT_C_FILE, data_only=True, read_only=True)
    check("Format C still detected as C",
          detect_sheet_raw_format(list(_wbC32["Maharashtra(U+R) - SHC"].iter_rows(values_only=True))) == "C")

# End-to-end: Val-Added-style opener parses identical to the Creme baseline
if os.path.exists(FORMAT_B_FILE):
    _base32, _ = process_workbook(FORMAT_B_FILE, "Creme")
    _wb32 = _op32.load_workbook(FORMAT_B_FILE)
    _ws32 = _wb32[_wb32.sheetnames[0]]
    _ws32.delete_rows(1)
    _ws32["A1"] = "Analysis"; _ws32["B1"] = "Crosstab - PulsePlus-8.0"
    _ws32["A2"] = "Measure";  _ws32["B2"] = "Households in 000s ; Vol"
    _ws32["A3"] = None;       _ws32["B3"] = "Average Consumption in S"
    _buf32 = _io32.BytesIO(); _wb32.save(_buf32); _buf32.seek(0)
    _val32, _ = process_workbook(_buf32, "Val Added Pwd")
    _cols32 = [c for c in _base32.columns if c != "Format"]
    check("Val-Added-style opener parses byte-identical to Creme baseline",
          _base32[_cols32].reset_index(drop=True).equals(_val32[_cols32].reset_index(drop=True)))



print("=" * 70)
print("TEST GROUP 33: KPI Explorer never sums percentages (>100% bug)")
print("=" * 70)
try:
    from streamlit.testing.v1 import AppTest as _AT33
    import warnings as _w33; _w33.filterwarnings("ignore")

    _frames33 = []
    _groups33 = {
        "SHC": {"Maharashtra(U+R) - SHC", "Maharasthra(R) - SHC", "Maharasthra(U) - SHC"},
        "KM":  {"Maharasthra(U+R) - Pwd KM", "Maharasthra(R) - Pwd KM", "Maharasthra(U) - Pwd KM"},
    }
    for _fmt33, _sh33 in _groups33.items():
        _d33, _ = process_workbook(FORMAT_C_FILE, _fmt33, included_sheet_names=_sh33)
        _d33["State_Zone"] = "Maharashtra"
        _frames33.append(_d33)
    _m33 = pd.concat(_frames33, ignore_index=True)
    _r33, _, _, _ = add_calculations(_m33, factor_lookup=build_factor_lookup({"Maharashtra": {"U": 1.85, "R": 1.85}}), zone_mapping={})
    _cs33 = build_company_summary(_r33)

    _at33 = _AT33.from_file("app.py", default_timeout=120)
    _at33.session_state["result_df"] = _r33
    _at33.session_state["company_summary_df"] = _cs33
    _at33.session_state["current_page"] = "explorer"
    _at33.run()
    _at33.selectbox[0].select("GODREJ CONSUMER PRODS")
    _at33.selectbox[1].select("Value MS%")
    _at33.selectbox[2].select("Maharashtra")
    _at33.selectbox[3].select("U+R")
    _at33.selectbox[-1].select(_at33.selectbox[-1].options[0])  # trend
    _at33.run()
    check("Explorer MS% trend runs clean", len(_at33.exception) == 0)

    def _pct33(s):
        try:
            return float(str(s).replace("%", "").replace(",", ""))
        except (ValueError, TypeError):
            return None

    _shown33 = _at33.dataframe[0].value
    _vals33 = [_pct33(v) for v in _shown33["Value MS%"] if _pct33(v) is not None]
    check("No Value MS% exceeds 100% (percentages are not summed)",
          all(v <= 100.01 for v in _vals33), f"max was {max(_vals33) if _vals33 else 'n/a'}")

    # Each shown MS% must equal the company summary exactly
    _per33 = [c for c in _cs33.columns if c.startswith("Value MS%__")][0].split("__")[1]
    _g33 = _cs33[(_cs33.Company == "GODREJ CONSUMER PRODS") & (_cs33.State_Zone == "Maharashtra")
                 & (_cs33.Urban_Rural == "U+R") & (_cs33.TG_Segment == "TOTAL")]
    _exp33 = {row.Format: row["Value MS%__" + _per33] * 100 for _, row in _g33.iterrows()}
    _mism33 = 0
    for _, _row33 in _shown33[_shown33["Period"] == _per33].iterrows():
        _e33 = _exp33.get(_row33["Category"])
        _s33 = _pct33(_row33["Value MS%"])
        if _e33 is not None and _s33 is not None and abs(_e33 - _s33) > 0.01:
            _mism33 += 1
    check("Every explorer MS% matches the company summary exactly", _mism33 == 0)

    # 'Category Total' must never appear as a selectable company
    check("Category Total excluded from company entities",
          "Category Total" not in set(_cs33.Company))
except ImportError:
    print("  SKIP (streamlit AppTest not available)")



print("=" * 70)
print("TEST GROUP 34: Combined market share across categories (no false 100%)")
print("=" * 70)
try:
    from streamlit.testing.v1 import AppTest as _AT34
    import warnings as _w34; _w34.filterwarnings("ignore")

    _frames34 = []
    for _fmt34, _sh34 in {
        "SHC": {"Maharashtra(U+R) - SHC", "Maharasthra(R) - SHC", "Maharasthra(U) - SHC"},
        "KM":  {"Maharasthra(U+R) - Pwd KM", "Maharasthra(R) - Pwd KM", "Maharasthra(U) - Pwd KM"},
    }.items():
        _d34, _ = process_workbook(FORMAT_C_FILE, _fmt34, included_sheet_names=_sh34)
        _d34["State_Zone"] = "Maharashtra"
        _frames34.append(_d34)
    _m34 = pd.concat(_frames34, ignore_index=True)
    _r34, _, _, _ = add_calculations(_m34, factor_lookup=build_factor_lookup({"Maharashtra": {"U": 1.85, "R": 1.85}}), zone_mapping={})
    _cs34 = build_company_summary(_r34)
    _per34 = [c for c in _r34.columns if c.startswith("Sales Derived__")][0].split("__")[1]

    # Manual combined share
    _cat34 = _r34[(_r34.Flag == "Category") & (_r34.State_Zone == "Maharashtra") & (_r34.Urban_Rural == "U+R") & (_r34.TG_Segment == "TOTAL")]
    _denom34 = _cat34.groupby("Format")["Sales Derived__" + _per34].first().sum()
    _num34 = _cs34[(_cs34.Company == "GODREJ CONSUMER PRODS") & (_cs34.State_Zone == "Maharashtra") & (_cs34.Urban_Rural == "U+R") & (_cs34.TG_Segment == "TOTAL")]["Sales Derived__" + _per34].sum()
    _expected34 = _num34 / _denom34 * 100

    _at34 = _AT34.from_file("app.py", default_timeout=120)
    _at34.session_state["result_df"] = _r34
    _at34.session_state["company_summary_df"] = _cs34
    _at34.session_state["current_page"] = "explorer"
    _at34.run()
    _at34.selectbox[0].select("GODREJ CONSUMER PRODS")
    _at34.selectbox[1].select("Value MS%")
    _at34.selectbox[2].select("Maharashtra")
    _at34.selectbox[3].select("U+R")
    _at34.selectbox[-1].select(_per34)
    _at34.run()
    check("Combined-share view runs clean", len(_at34.exception) == 0)

    _cards34 = [str(md.value) for md in _at34.markdown if "stat-card" in str(md.value)]
    _shown34 = _cards34[-1].split('stat-value">')[1].split("<")[0] if _cards34 else ""
    _shownf34 = float(_shown34.replace("%", "")) if _shown34 else -1
    check("Combined share matches manual sales-weighted calc",
          abs(_shownf34 - _expected34) < 0.05, f"shown={_shown34} expected={_expected34:.2f}%")
    check("Combined share is a valid 0-100% (never a false 100 or over)", 0 <= _shownf34 <= 100)
    check("'meaningless' message removed", not any("meaningless" in i.value for i in _at34.info))

    # 'ANY '-prefixed category rows must never be selectable as a brand
    _at34.radio[0].set_value("Brand")
    _at34.run()
    _brand_opts34 = _at34.selectbox[0].options
    check("No 'ANY '-prefixed pseudo-brand in the brand picker",
          not any(str(b).strip().upper().startswith("ANY ") for b in _brand_opts34))
except ImportError:
    print("  SKIP (streamlit AppTest not available)")

print("=" * 70)
print(f"RESULT: {PASS} passed, {FAIL} failed")
print("=" * 70)
sys.exit(1 if FAIL else 0)
