"""
Godrej Market Share Model - Data Cleaning Module
Reads raw Kantar 68-sheet Excel files (one per Format/Category) and produces
a single clean long-form dataframe: one row per Region x TG_Segment x Brand x Period.

Key facts validated against ground truth before writing this code:
- Column A (TG) in the raw sheet is filled only on real data rows, with the
  segment name (TOTAL, SEC A, SEC B, SEC C, SEC D/E). Blank / header / Universe
  rows always have column A empty. This lets us detect segments and skip
  junk rows without guessing row numbers.
- The metric header row (HH, Vol, Val, Avg Cons, Avg FOP, Avg POC, Avg NOP) and
  the period header row are identical everywhere in the sheet, so we only need
  to read them once (row 4 and row 5) to build column names.
- The "[X] ANY ..." row appearing first in every segment is the Category total.
  Any other bracketed row ("[HCEXL] ..." or "[COLOR] ...") is a Brand row.
  Un-bracketed rows are individual SKUs and are NOT needed for this model
  (Sagar's brand-level share model only needs Category and Brand rows; Brand
  totals from Kantar already include unlisted small SKUs, so summing children
  would double count / under count).
"""

import openpyxl
import pandas as pd
import re
from default_brand_mapping import DEFAULT_BRAND_MAPPING
from default_company_mapping import DEFAULT_COMPANY_MAPPING

TG_SEGMENTS = {"TOTAL", "SEC A", "SEC B", "SEC C", "SEC D/E"}

# Sheets to always exclude (per Sagar's rule: keep "Andhra Pradesh excl Tela",
# drop any combined "AP+Tel" sheet to avoid double counting)
EXCLUDE_SHEET_KEYWORDS = ["AP+Tel", "AP + Tel", "AP&Tel"]

# Known zone identifiers. A sheet base-name is treated as a Zone (not a
# State) if it matches one of these, or contains "zone"/"gcpl". This matters
# because real zone sheets are often just named "North", "GCPL East" etc,
# with no literal word "Zone" in them - a name-pattern-only check would
# misclassify them as ordinary states.
KNOWN_ZONE_NAMES = {"north", "south", "east", "west"}


def parse_region_sheet_name(sheet_name):
    """
    Split a sheet name into (State/Zone name, Urban_Rural, is_zone).
    Handles patterns like:
      'Rajasthan (U+R)', 'Rajasthan (U)', 'Rajasthan (R)'
      'All India Urban', 'All India Rural', 'All India U+R'
      'North (U)', 'North Zone(U)', 'GCPL East (U+R)'
      'Andhra Pradesh excl Tela (U+R)'
    """
    name = sheet_name.strip()

    # All India special case (no parentheses, word Urban/Rural/U+R at the end)
    m = re.match(r"^All India\s*(Urban|Rural|U\+R)$", name, re.IGNORECASE)
    if m:
        ur_raw = m.group(1)
        ur = {"urban": "U", "rural": "R", "u+r": "U+R"}[ur_raw.lower()]
        return "All India", ur, False

    # Generic pattern: 'BaseName (U+R)' / 'BaseName(U)' etc - covers both
    # states ('Rajasthan') and zones ('North', 'GCPL East', 'North Zone').
    # No end-of-string anchor: some sheet names have extra text after the
    # UR marker (e.g. Format C's combined "Maharashtra(U+R) -Val added pwd"
    # sample sheets), which is ignored here since Format is assigned
    # separately by the person uploading the file, not parsed from the name.
    m = re.match(r"^(.*?)\s*\((U\+R|U|R)\)", name)
    if m:
        base = m.group(1).strip()
        ur = m.group(2).upper()
        base_clean = re.sub(r"\s*Zone\s*$", "", base, flags=re.IGNORECASE).strip()
        is_zone = (
            base_clean.lower() in KNOWN_ZONE_NAMES
            or "zone" in base.lower()
            or "gcpl" in base.lower()
        )
        return base_clean, ur, is_zone

    # Fallback: couldn't parse, return as-is
    return name, "UNKNOWN", False


def is_excluded_sheet(sheet_name, exclude_keywords=None):
    keywords = exclude_keywords if exclude_keywords is not None else EXCLUDE_SHEET_KEYWORDS
    return any(kw.lower() in sheet_name.lower() for kw in keywords if kw)


def get_sheet_overview(file_path_or_bytes, exclude_keywords=None):
    """
    Lists every sheet in a raw workbook with its classification, without
    reading any row data (cheap - just sheet names). Used by the Region
    Selection page so the person can see and choose exactly which sheets
    get processed, instead of a hidden keyword rule.

    Default suggested inclusion:
      - Plain State sheets -> included
      - Zone sheets (North, GCPL East, etc) -> excluded (zones are
        calculated by summing states via Zone Mapping, not from their own
        sheet - see calculator.add_calculations)
      - AP+Tel-style combined sheets -> excluded (avoids double counting
        with Andhra Pradesh excl Tela + Telangana reported separately)
      - "All India" sheets -> excluded, for exactly the same reason as Zone
        sheets: calculator.add_calculations always rebuilds All India as the
        sum of every real state present and explicitly discards any raw
        "All India" rows, so reading that sheet is wasted work whose result
        is thrown away. Leaving it ticked also risks a hard failure on files
        (e.g. Henna) whose All India sheet uses a different layout to the
        state sheets, for no benefit whatsoever. It is still listed here and
        can be ticked manually, but it will never affect the numbers.
    """
    if hasattr(file_path_or_bytes, "seek"):
        file_path_or_bytes.seek(0)
    wb = openpyxl.load_workbook(file_path_or_bytes, read_only=True)
    overview = []
    for sheet_name in wb.sheetnames:
        state_or_zone, urban_rural, is_zone = parse_region_sheet_name(sheet_name)
        is_ap_tel = is_excluded_sheet(sheet_name, exclude_keywords=exclude_keywords)
        is_all_india = str(state_or_zone).strip().lower() == "all india"
        default_include = (not is_zone) and (not is_ap_tel) and (not is_all_india)
        overview.append({
            "Sheet_Name": sheet_name,
            "State_Zone": state_or_zone,
            "Urban_Rural": urban_rural,
            "Is_Zone": is_zone,
            "Is_All_India": is_all_india,
            "Is_AP_Tel": is_ap_tel,
            "Include": default_include,
        })
    wb.close()
    return overview


def find_header_rows(rows, id_cols=5, max_scan=20):
    """
    Dynamically locates the metric-label row (the row containing 'HH', the
    first of the HH/Vol/Val/... metric blocks) and treats the very next row
    as the period-label row. This replaces a previous fixed-row-position
    assumption (always row 4 / row 5) which broke silently - producing
    garbage column names - whenever a raw file had a slightly different
    header layout (an extra or missing row above the data).
    Returns (metric_header_row_idx, period_header_row_idx) or (None, None)
    if no 'HH' label is found within the scanned rows.
    """
    for i in range(min(max_scan, len(rows))):
        row = rows[i]
        for cell in row[id_cols:id_cols + 3]:
            if cell is not None and str(cell).strip().upper() == "HH":
                return i, i + 1
    return None, None


def build_column_map(rows, id_cols=5):
    """
    Build a list of column names for columns starting at index `id_cols`.
    rows: list of row tuples (0-indexed, as returned by ws.iter_rows(values_only=True))
    Returns list of strings like 'HH__2023 Apr - 2023 Jun'.

    Raises ValueError (rather than silently producing wrong data) if the
    header rows can't be located, or if what's found at the expected period
    position doesn't look like a real period label (e.g. is a bare number,
    which means the sheet's layout doesn't match what was detected and
    something has shifted).
    """
    metric_header_row_idx, period_header_row_idx = find_header_rows(rows, id_cols=id_cols)
    if metric_header_row_idx is None:
        raise ValueError(
            "Could not find the 'HH' metric header row in this sheet within the first "
            "20 rows. This sheet's layout doesn't match the expected Kantar export "
            "structure (TG/Format/Grammage/SU/Product columns followed by HH/Vol/Val/... "
            "metric blocks). Nothing was read from this sheet rather than risk reading "
            "the wrong cells."
        )

    metric_row = rows[metric_header_row_idx]
    period_row = rows[period_header_row_idx]

    col_names = []
    current_metric = None
    for i in range(id_cols, len(metric_row)):
        if metric_row[i] is not None:
            current_metric = str(metric_row[i]).strip()
        period = period_row[i]
        if period is None:
            col_names.append(None)
            continue
        period_str = str(period).strip()
        try:
            float(period_str)
            looks_numeric = True
        except ValueError:
            looks_numeric = False
        if looks_numeric or current_metric is None:
            raise ValueError(
                f"Header row looks wrong at column index {i}: metric='{current_metric}', "
                f"period='{period_str}' (a period label should be text like '2023 Apr - "
                f"2023 Jun', not a number). This means the detected header row doesn't "
                f"match this sheet's real layout. Check this file's structure around rows "
                f"{metric_header_row_idx + 1}-{period_header_row_idx + 1}."
            )
        col_names.append(f"{current_metric}__{period_str}")
    return col_names


def classify_flag(product_name, flag_overrides=None):
    """
    Returns one of: 'Category', 'Brand', 'Sub-brand', 'Others', 'SKU'

    Priority order:
      1. flag_overrides (the Brand Mapping table) - checked FIRST, regardless
         of bracket notation. This matters because not every format uses the
         same naming convention (e.g. HBP/Powder has at least one real Brand
         row, "5 Star", with no brackets at all - the mapping catches this,
         a pure bracket-based guess would not).
      2. '[X] ANY ...' pattern -> Category (one per segment, the format total)
      3. Any other bracketed row -> Others (if it contains "OTH.") else Brand
      4. Un-bracketed and not in the mapping -> SKU (excluded from output,
         since Sagar's brand-level share model doesn't need individual SKUs)

    flag_overrides: dict {Brand_SKU_Item (exact text): 'Category'|'Brand'|'Sub-brand'|'Others'}
    """
    name = str(product_name).strip()

    if flag_overrides and name in flag_overrides:
        return flag_overrides[name]

    if not name.startswith("["):
        return "SKU"

    after_bracket = re.sub(r"^\[[^\]]+\]\s*", "", name)
    if after_bracket.upper().startswith("ANY "):
        return "Category"

    if "OTH." in name.upper() or "OTHERS" in name.upper():
        return "Others"

    return "Brand"


def classify_company(product_name, flag, company_overrides=None):
    """
    Returns the parent Company for a Brand_SKU_Item, e.g. 'GODREJ CONSUMER PRODS'.
    Category rows get 'Category Total'. Others/unmapped items not found in the
    mapping get 'Others / Unmapped' rather than being silently blank, so they're
    still visible to review in the output.

    company_overrides: dict {Brand_SKU_Item (exact text): Company name}
    """
    if flag == "Category":
        return "Category Total"

    name = str(product_name).strip()
    if company_overrides and name in company_overrides:
        return company_overrides[name]

    return "Others / Unmapped"


_ANY_SUFFIX_PATTERN = re.compile(r"\s*\[ANY[^\]]*\]\s*$", re.IGNORECASE)


def _strip_any_suffix(name):
    """
    Strips a trailing '[ANY <category>]' suffix if present, e.g.
    '[HCEXL] ABIHA POWDER [ANY HERBAL BASED PWD]' -> '[HCEXL] ABIHA POWDER'.
    Some raw formats append this redundant category suffix to every brand
    name (found in the older ground-truth naming for HBP specifically);
    newer formats (e.g. Format C) do not. Rather than requiring two
    separately-maintained mapping files, the lookup table is extended with
    the suffix-stripped form as a fallback, so the same underlying brand is
    recognized either way.
    """
    return _ANY_SUFFIX_PATTERN.sub("", str(name)).strip()


_BRACKET_PREFIX_PATTERN = re.compile(r"^\[[^\]]+\]\s*")


def _strip_bracket_prefix(name):
    """
    Strips a leading '[TAG] ' bracket prefix if present, e.g.
    '[HCEXL] GODREJ PROF.DIMENSION CREAM HAIR COLOR' -> 'GODREJ
    PROF.DIMENSION CREAM HAIR COLOR'. Format B's raw files never use bracket
    tags at all, so matching against the reference mapping (which does use
    them) needs this fallback the same way HBP needed the suffix fallback.
    """
    return _BRACKET_PREFIX_PATTERN.sub("", str(name)).strip()


def _with_naming_fallbacks(mapping):
    """
    Returns a copy of `mapping` extended with both suffix-stripped (e.g.
    HBP's redundant '[ANY category]' suffix) and prefix-stripped (e.g.
    Format B's missing '[TAG] ' bracket prefix) fallback keys, plus both
    combined. Never overwrites an existing distinct entry, to avoid
    silently masking a real difference.
    """
    extended = dict(mapping)
    for key, value in mapping.items():
        candidates = {
            _strip_any_suffix(key),
            _strip_bracket_prefix(key),
            _strip_bracket_prefix(_strip_any_suffix(key)),
        }
        for candidate in candidates:
            if candidate != key and candidate not in extended:
                extended[candidate] = value
    return extended


def _with_suffix_fallback(mapping):
    """Deprecated alias, kept for compatibility - use _with_naming_fallbacks."""
    return _with_naming_fallbacks(mapping)


_QUARTER_SEASON_MAP = {"AMJ": ("Apr", "Jun"), "JAS": ("Jul", "Sep"), "OND": ("Oct", "Dec"), "JFM": ("Jan", "Mar")}


def _format_c_period_to_canonical(period_code):
    """
    Converts a Format C period code, e.g. "01. AMJ'23" or "MAT Mar24", into
    the exact same canonical period label Format A already uses (e.g.
    "2023 Apr - 2023 Jun" or "2023 Apr - 2024 Mar"). This means every
    downstream module (calculator.py, exporter.py) never needs to know or
    care which raw format a file came from - both formats produce the same
    column names once cleaned.
    Returns the canonical label, or None if the code doesn't match either
    the quarterly or MAT pattern.
    """
    code = str(period_code).strip()
    m_mat = re.match(r"^MAT\s+[A-Za-z]{3}(\d{2})$", code, re.IGNORECASE)
    if m_mat:
        end_year = 2000 + int(m_mat.group(1))
        start_year = end_year - 1
        return f"{start_year} Apr - {end_year} Mar"

    m_q = re.match(r"^\d+\.\s*([A-Za-z]{3})'(\d{2})$", code)
    if m_q:
        season, yy = m_q.group(1).upper(), m_q.group(2)
        if season in _QUARTER_SEASON_MAP:
            start_m, end_m = _QUARTER_SEASON_MAP[season]
            year = 2000 + int(yy)
            return f"{year} {start_m} - {year} {end_m}"
    return None


def find_format_c_header(rows, max_scan=15):
    """
    Format C sheets have a single flat header row (position varies - row 2 or
    row 3 depending on the sheet) containing a 'Brand_SKU Item' cell, and
    combined metric+period column names like "HH_01. AMJ'23" or
    "HH_MAT Mar24". This finds that row and how many identifier columns
    precede the metric columns (5 for most formats, 6 for HBP/Henna which
    have one extra column before Brand_SKU Item).
    Returns (header_row_idx, id_cols) or (None, None) if not found.
    """
    for i in range(min(max_scan, len(rows))):
        row = rows[i]
        for j, cell in enumerate(row):
            if cell is not None and "Brand_SKU" in str(cell):
                return i, j + 1
    return None, None


def build_column_map_format_c(rows, header_row_idx, id_cols):
    """
    Format C equivalent of build_column_map: builds canonical
    'Metric__Period' column names from a single flat header row, instead of
    two separate rows. Raises ValueError (rather than silently producing
    wrong data) if a header cell doesn't match the expected pattern.
    """
    header = rows[header_row_idx]
    col_names = []
    for i in range(id_cols, len(header)):
        cell = header[i]
        if cell is None:
            col_names.append(None)
            continue
        cell_str = str(cell).strip()
        if "_" not in cell_str:
            raise ValueError(
                f"Header cell '{cell_str}' at column index {i} doesn't look like a Format C "
                f"metric column (expected something like \"HH_01. AMJ'23\")."
            )
        metric, period_code = cell_str.split("_", 1)
        canonical_period = _format_c_period_to_canonical(period_code)
        if canonical_period is None:
            raise ValueError(
                f"Could not parse period code '{period_code}' from header cell '{cell_str}' "
                f"at column index {i}."
            )
        col_names.append(f"{metric.strip()}__{canonical_period}")
    return col_names


def classify_flag_format_b(product_name):
    """
    Format B has no bracket-tag naming convention at all. Category and Brand
    rows are instead distinguished by a leading AND trailing space around the
    name (e.g. ' ANY CREME ', ' MATRIX SOCOLOR PREBONDED CREAM HAIR COLO ');
    SKU rows have no surrounding space. Confirmed directly against a real
    Format B file, not assumed.
    """
    raw = str(product_name)
    if raw.startswith(" ") and raw.endswith(" "):
        trimmed = raw.strip()
        upper = trimmed.upper()
        if upper.startswith("ANY "):
            # Most 'ANY ...' rows are the genuine category total (e.g.
            # 'ANY HAIR COL.SHAMPOO'). BUT some files also contain an
            # 'ANY <company> ...' subtotal - e.g. 'ANY GCPL SHC (GEE+SELFIE)'
            # is Godrej's own Expert+Selfie subtotal, NOT the market total.
            # If such a subtotal is treated as the Category denominator, the
            # company's MS% collapses to its own value / its own value = 100%
            # (this is exactly the SHC 100% bug). So an 'ANY' row that names a
            # specific manufacturer/brand is classified as a Brand subtotal,
            # never as the category. The true category total names the product,
            # not a company.
            COMPANY_TOKENS_IN_ANY = (
                "GCPL", "GODREJ", "HUL", "LOREAL", "L'OREAL", "CAVINKARE",
                "HENKEL", "MARICO", "PATANJALI", "HONASA", "GEE", "SELFIE",
                "EXPERT", "HYGIENIC", "VCARE", "SISO", "NEHA",
            )
            if any(tok in upper for tok in COMPANY_TOKENS_IN_ANY):
                return "Subtotal"
            return "Category"
        if "OTH." in upper or "OTHERS" in upper:
            return "Others"
        return "Brand"
    return "SKU"


def find_format_b_signature(rows, max_scan=6):
    """
    Format B sheets open with a small header block that is unique to this
    format and present in neither Format A nor Format C (both of which start
    with a 'TG'/'Customer Segment' + 'Format' + 'Grammage' style column
    header instead). Checked first so it can never be confused with them.

    Two accepted variants of that block, both seen in real Kantar files:

      1. A 'Process Period' row (e.g. 'Process Period' | '2025 Jun To
         2026 May') - the Creme-style files.
      2. NO 'Process Period' row at all, starting directly with the
         'Universe' | 'With Growth Factor' row - the Henna-style files.
         Structurally these are identical to variant 1 with the first row
         removed, so they parse through exactly the same path; only the
         signature row differs.

    Matching on 'Universe' alone would be too loose, so variant 2 requires
    the 'With Growth Factor' partner cell beside it.
    """
    for i in range(min(max_scan, len(rows))):
        row = rows[i]
        if not row or row[0] is None:
            continue
        first = str(row[0]).strip().lower()
        if first == "process period":
            return True
        if first == "universe" and len(row) > 1 and row[1] is not None:
            if "growth factor" in str(row[1]).strip().lower():
                return True
        # Variant 3: some files (e.g. Val Added Pwd) open with an
        # 'Analysis' | 'Crosstab - PulsePlus-8.0' row instead of Process
        # Period or Universe. Everything below that opener is the same
        # Format B monthly block. Matched on the 'crosstab' + 'pulseplus'
        # pair (Kantar's product name) so it stays specific - neither word
        # appears at the top of Format A or Format C.
        if first == "analysis" and len(row) > 1 and row[1] is not None:
            partner = str(row[1]).strip().lower()
            if "crosstab" in partner and "pulseplus" in partner.replace(" ", "").replace("-", ""):
                return True
    return False


def detect_sheet_raw_format(rows):
    """
    Returns 'B' if this sheet matches the newer monthly-only format (starts
    with a 'Process Period' marker), 'C' if it matches the flat-header
    format (has a 'Brand_SKU Item' column), 'A' if it matches the original
    two-row-header format (has a standalone 'HH' label row), or None if none
    of the three is recognized.
    """
    if find_format_b_signature(rows):
        return "B"
    header_idx, _ = find_format_c_header(rows)
    if header_idx is not None:
        return "C"
    metric_idx, _ = find_header_rows(rows)
    if metric_idx is not None:
        return "A"
    return None


_SINGLE_MONTH_PATTERN = re.compile(r"^\d{4} [A-Za-z]{3}$")


_MONTH_NUM = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
              "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}


def _month_period_sort_key(period_label):
    """Sorts a 'YYYY Mon' label in true calendar order, not alphabetically
    (alphabetical sort would wrongly put '2025 Aug' before '2025 Jun')."""
    year_str, mon_str = period_label.split(" ")
    return (int(year_str), _MONTH_NUM.get(mon_str, 0))


def add_format_b_annual_totals(df, format_tag=None, include_estimated_avg_nop=False):
    """
    Format B files have no pre-computed MAT (annual) columns, only 12
    individual months. This computes an annual total for the metrics that
    are validly additive across months (HH, Vol, Val - true monthly flow
    quantities, the same way Kantar's own MAT works), labeled honestly with
    the file's own actual date range (e.g. "2025 Jun - 2026 May") rather
    than a fixed "MAT Mar24"-style label, since the window may not align to
    the standard Apr-Mar fiscal year Sagar's other benchmarks use.

    Only ever sums genuine single-month columns (period label matching
    "YYYY Mon", e.g. "2025 Jun") - this makes it safe to call even on a
    combined table with Format A/C data mixed in, since their quarterly
    ("2023 Apr - 2023 Jun") and MAT ("2023 Apr - 2024 Mar") columns simply
    don't match that pattern and are left untouched, never double-summed.

    Deliberately does NOT compute an annual figure for Avg Cons / Avg POC /
    Avg NOP under the standard column name: these are average-type metrics,
    and summing 12 monthly averages would produce a meaningless inflated
    number. A valid annual average would need to be derived as a ratio of
    annual totals (e.g. Annual Vol / Annual HH), but the exact intended
    definition hasn't been confirmed. If include_estimated_avg_nop=True, a
    best-effort estimate (Annual Vol / Annual HH) is added under a clearly
    marked column name ("Avg NOP (ESTIMATED)__...") rather than the exact
    name calculator.py treats as authoritative, so it can be inspected but
    is never silently used to compute Units Estd until confirmed correct.
    """
    if format_tag is not None:
        mask = df["Format"] == format_tag
        if not mask.any():
            return df
    else:
        mask = pd.Series(True, index=df.index)

    df = df.copy()
    annual_col_by_metric = {}
    for metric in ["HH", "Vol", "Val"]:
        cols = [
            c for c in df.columns
            if c.startswith(f"{metric}__") and _SINGLE_MONTH_PATTERN.match(c.split("__", 1)[1])
        ]
        cols = sorted(cols, key=lambda c: _month_period_sort_key(c.split("__", 1)[1]))
        if len(cols) < 2:
            continue
        first_period = cols[0].split("__", 1)[1]
        last_period = cols[-1].split("__", 1)[1]
        annual_col = f"{metric}__{first_period} - {last_period}"
        df.loc[mask, annual_col] = df.loc[mask, cols].sum(axis=1, min_count=1)
        annual_col_by_metric[metric] = annual_col

    if include_estimated_avg_nop and "Vol" in annual_col_by_metric and "HH" in annual_col_by_metric:
        vol_col = annual_col_by_metric["Vol"]
        hh_col = annual_col_by_metric["HH"]
        period_label = vol_col.split("__", 1)[1]
        est_col = f"Avg NOP (ESTIMATED)__{period_label}"
        hh_vals = df.loc[mask, hh_col]
        vol_vals = df.loc[mask, vol_col]
        df.loc[mask, est_col] = vol_vals.where(hh_vals != 0) / hh_vals.replace(0, pd.NA)
    return df


def process_workbook(file_path_or_bytes, format_tag, flag_overrides=None, company_overrides=None, exclude_keywords=None, included_sheet_names=None, verbose=False):
    """
    Reads one raw Kantar workbook (all sheets = all regions) for a single
    Format/Category, and returns (df, sku_df):
      df     - Category/Brand/Sub-brand/Others rows (the main Master_Clean data)
      sku_df - individual SKU rows, each tagged with its Parent_Brand, kept
               separately (not part of Master_Clean) purely so a Variance
               check can compare a Brand's own reported total against the
               sum of its listed SKUs.

    flag_overrides (if given) are layered on top of DEFAULT_BRAND_MAPPING,
    i.e. user edits/uploads always win over the built-in default.

    included_sheet_names: if given (a set/list of exact sheet names), ONLY
    those sheets are processed - this is what the Region Selection page
    drives. If None, falls back to the old keyword-based exclude_keywords
    rule (kept for programmatic/test use).
    """
    combined_mapping = dict(DEFAULT_BRAND_MAPPING)
    if flag_overrides:
        combined_mapping.update(flag_overrides)
    combined_mapping = _with_suffix_fallback(combined_mapping)

    combined_company_mapping = dict(DEFAULT_COMPANY_MAPPING)
    if company_overrides:
        combined_company_mapping.update(company_overrides)
    combined_company_mapping = _with_suffix_fallback(combined_company_mapping)

    if hasattr(file_path_or_bytes, "seek"):
        file_path_or_bytes.seek(0)
    wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True, read_only=True)

    all_records = []
    sku_records = []
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        if included_sheet_names is not None:
            if sheet_name not in included_sheet_names:
                skipped_sheets.append(sheet_name)
                continue
        elif is_excluded_sheet(sheet_name, exclude_keywords=exclude_keywords):
            skipped_sheets.append(sheet_name)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue

        state_or_zone, urban_rural, is_zone = parse_region_sheet_name(sheet_name)

        raw_format = detect_sheet_raw_format(rows)
        if raw_format == "B":
            metric_header_row_idx, period_header_row_idx = find_header_rows(rows, id_cols=1)
            if metric_header_row_idx is None:
                raise ValueError(
                    f"[{format_tag} / sheet '{sheet_name}'] Detected as Format B ('Process Period' "
                    f"marker found) but could not locate the 'HH' metric header row. Nothing was "
                    f"read from this sheet rather than risk reading the wrong cells."
                )
            try:
                col_names = build_column_map(rows, id_cols=1)
            except ValueError as e:
                raise ValueError(f"[{format_tag} / sheet '{sheet_name}'] {e}") from e
            id_cols = 1
            data_start_idx = period_header_row_idx + 1
        elif raw_format == "C":
            header_row_idx, id_cols = find_format_c_header(rows)
            try:
                col_names = build_column_map_format_c(rows, header_row_idx, id_cols)
            except ValueError as e:
                raise ValueError(f"[{format_tag} / sheet '{sheet_name}'] {e}") from e
            data_start_idx = header_row_idx + 1
        elif raw_format == "A":
            try:
                col_names = build_column_map(rows)
            except ValueError as e:
                raise ValueError(f"[{format_tag} / sheet '{sheet_name}'] {e}") from e
            metric_header_row_idx, period_header_row_idx = find_header_rows(rows)
            id_cols = 5
            data_start_idx = period_header_row_idx + 1
        else:
            seen = []
            for _r in rows[:6]:
                if not _r:
                    continue
                _a = str(_r[0]).strip() if _r[0] is not None else ""
                _b = str(_r[1]).strip() if len(_r) > 1 and _r[1] is not None else ""
                if _a or _b:
                    seen.append(f"'{_a[:24]}' | '{_b[:24]}'")
            seen_txt = " ;; ".join(seen[:4]) if seen else "(first 6 rows are all empty)"
            raise ValueError(
                f"[{format_tag} / sheet '{sheet_name}'] Could not recognize this sheet's layout as "
                f"any known format (Format A: standalone 'HH' header row with 5 id columns; Format B: "
                f"a 'Process Period' row OR a 'Universe' + 'With Growth Factor' row; Format C: a "
                f"'Brand_SKU Item' column). Nothing was read from this sheet rather than risk reading "
                f"the wrong cells. First rows actually seen (col A | col B): {seen_txt}"
            )

        product_idx = id_cols - 1
        metrics_start = id_cols

        current_parent_brand = None

        for r in rows[data_start_idx:]:  # data starts right after the header row; non-data rows (Universe, TG Base, blanks) are filtered out below by the TG_SEGMENTS check
            if raw_format == "B":
                # Format B has no TG_Segment column at all - every row is
                # implicitly TOTAL (confirmed directly against a real file:
                # zero SEC A/B/C/D markers found anywhere in 519 data rows).
                tg_segment = "TOTAL"
                product = r[0]
            else:
                tg_segment = r[0]
                product = r[product_idx]
            if tg_segment is None or product is None:
                continue  # junk / blank / universe / TG base row
            if tg_segment not in TG_SEGMENTS:
                continue

            if raw_format == "B":
                flag = classify_flag_format_b(product)
                if combined_mapping.get(str(product).strip()):
                    flag = combined_mapping[str(product).strip()]  # explicit override still wins
            else:
                flag = classify_flag(product, flag_overrides=combined_mapping)

            if flag == "SKU":
                if current_parent_brand is not None:
                    sku_record = {
                        "Format": format_tag,
                        "State_Zone": state_or_zone,
                        "Is_Zone": is_zone,
                        "Urban_Rural": urban_rural,
                        "TG_Segment": tg_segment,
                        "Parent_Brand": current_parent_brand,
                        "Brand_SKU_Item": str(product).strip(),
                    }
                    for i, col_name in enumerate(col_names):
                        if col_name is None:
                            continue
                        sku_record[col_name] = r[metrics_start + i]
                    sku_records.append(sku_record)
                continue  # SKU rows are never part of Master_Clean

            # This is a Category/Brand/Sub-brand/Others row - it becomes the
            # parent for any SKU rows that follow it, until the next one.
            current_parent_brand = str(product).strip()

            company = classify_company(product, flag, company_overrides=combined_company_mapping)

            record = {
                "Format": format_tag,
                "Region_Raw": sheet_name,
                "State_Zone": state_or_zone,
                "Is_Zone": is_zone,
                "Urban_Rural": urban_rural,
                "TG_Segment": tg_segment,
                "Flag": flag,
                "Company": company,
                "Grammage": r[2] if id_cols >= 4 else None,
                "SU": r[3] if id_cols >= 4 else None,
                "Brand_SKU_Item": str(product).strip(),
            }
            for i, col_name in enumerate(col_names):
                if col_name is None:
                    continue
                val = r[metrics_start + i]
                record[col_name] = val

            all_records.append(record)

    df = pd.DataFrame(all_records)
    sku_df = pd.DataFrame(sku_records)
    if verbose:
        print(f"[{format_tag}] Sheets read: {len(wb.sheetnames) - len(skipped_sheets)}, "
              f"skipped (AP+Tel): {skipped_sheets}, rows extracted: {len(df)}, SKU rows: {len(sku_df)}")
    return df, sku_df


def process_all_files(file_dict, flag_overrides=None, company_overrides=None, exclude_keywords=None, included_sheets_by_format=None, verbose=False):
    """
    file_dict: {format_tag: file_path_or_bytes}
    included_sheets_by_format: optional {format_tag: set of exact sheet names to include}
    Returns (master_df, sku_df) - the combined main data and the combined
    SKU-level data (used only for the Variance check), across all formats.
    """
    frames = []
    sku_frames = []
    for format_tag, f in file_dict.items():
        included = included_sheets_by_format.get(format_tag) if included_sheets_by_format else None
        df, sku_df = process_workbook(
            f, format_tag, flag_overrides=flag_overrides,
            company_overrides=company_overrides, exclude_keywords=exclude_keywords,
            included_sheet_names=included, verbose=verbose,
        )
        frames.append(df)
        sku_frames.append(sku_df)
    master_df = pd.concat(frames, ignore_index=True)
    sku_master_df = pd.concat(sku_frames, ignore_index=True) if sku_frames else pd.DataFrame()
    return master_df, sku_master_df
